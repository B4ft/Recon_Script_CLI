#!/usr/bin/env python3
"""
External Network Penetration Testing Reconnaissance Framework
A modular reconnaissance tool for authorized penetration testing engagements
OPTIMIZED VERSION with batch processing, parallel execution, IP/Domain handling, and cross-platform support
"""

import argparse
import subprocess
import json
import sys
import os
from datetime import datetime
from typing import List, Dict, Set
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import platform
from pathlib import Path

class PlatformDetector:
    """Detects the current platform (Windows/WSL2/Linux/macOS)"""
    
    @staticmethod
    def get_platform():
        """Returns: 'windows', 'wsl2', 'linux', or 'darwin'"""
        system = platform.system().lower()
        
        if system == "linux":
            try:
                with open("/proc/version", "r") as f:
                    version = f.read().lower()
                    if "microsoft" in version or "wsl" in version:
                        return "wsl2"
            except:
                pass
            return "linux"
        
        elif system == "darwin":
            return "darwin"
        
        elif system == "windows":
            return "windows"
        
        return "unknown"
    
    @staticmethod
    def is_wsl2():
        return PlatformDetector.get_platform() == "wsl2"
    
    @staticmethod
    def is_native_windows():
        return PlatformDetector.get_platform() == "windows"


class DependencyManager:
    """Manages installation of all required tools"""
    
    def __init__(self):
        self.home = Path.home()
        self.go_path = self.home / "go" / "bin"
        self.platform = PlatformDetector.get_platform()
        
    def check_and_install_all(self):
        """Check and install all dependencies"""
        
        if self.platform == "windows":
            self.handle_native_windows()
            return
        
        print("ð Checking dependencies...")
        
        self.ensure_python_deps()
        
        if not self.check_command("go"):
            self.install_go()
        
        os.environ["PATH"] = f"{self.go_path}:{os.environ['PATH']}"
        
        self.ensure_go_tool("subfinder", "github.com/projectdiscovery/subfinder/v2/cmd/subfinder@latest")
        self.ensure_go_tool("dnsx", "github.com/projectdiscovery/dnsx/cmd/dnsx@latest")
        self.ensure_go_tool("naabu", "github.com/projectdiscovery/naabu/v2/cmd/naabu@latest")
        self.ensure_go_tool("hakrawler", "github.com/hakluke/hakrawler@latest")
        self.ensure_go_tool("cdncheck","github.com/projectdiscovery/cdncheck/cmd/cdncheck@latest")
        self.ensure_go_tool("github-subdomains", "github.com/gwen001/github-subdomains@latest")
        self.ensure_go_tool("gau",              "github.com/lc/gau/v2/cmd/gau@latest")
        self.ensure_go_tool("alterx",           "github.com/projectdiscovery/alterx/cmd/alterx@latest")
        self.ensure_go_tool("shuffledns",       "github.com/projectdiscovery/shuffledns/cmd/shuffledns@latest")
        self.ensure_go_tool("assetfinder",      "github.com/tomnomnom/assetfinder@latest")
        self.ensure_go_tool("shosubgo",          "github.com/incogbyte/shosubgo@latest")
        self.ensure_go_tool("asnmap",            "github.com/projectdiscovery/asnmap/cmd/asnmap@latest")
        self.ensure_pd_httpx()                                 # PD httpx (avoids python httpx conflict)
        self.ensure_python_tool("dnsgen",        "dnsgen")         # pip install dnsgen
        self.ensure_python_tool("sublist3r",     "sublist3r")      # pip install sublist3r
        self.ensure_365doms()                                       # fetch gist script

        self.check_system_tool("nmap")
        self.check_system_tool("dig")
        self.check_system_tool("whois")
        
        print("✅ All dependencies ready!\n")
    
    def handle_native_windows(self):
        print("\n" + "="*60)
        print("⚠️  WINDOWS DETECTED")
        print("="*60)
        print("\nFor best experience and full features, use WSL2 (Ubuntu).")
        print("\nQuick WSL2 Setup (5 minutes):")
        print("  1. Open PowerShell as Administrator")
        print("  2. Run: wsl --install Ubuntu")
        print("  3. Restart your computer")
        print("  4. Open Ubuntu terminal")
        print("  5. Clone the repo and run this script")
        print("\nWSL2 Benefits:")
        print("  ✅ Full feature support (port scanning, all tools)")
        print("  ✅ Better performance")
        print("  ✅ Automatic tool installation")
        print("  ✅ Native Linux environment")
        print("\n" + "="*60 + "\n")
        
        response = input("Continue with limited features on Windows? [y/N]: ")
        
        if response.lower() not in ['y', 'yes']:
            print("\nSetup instructions: https://docs.microsoft.com/en-us/windows/wsl/install")
            sys.exit(0)
        
        print("\n⚠️  Running in limited mode:")
        print("  ✅ Subdomain enumeration (if tools available)")
        print("  ✅ DNS resolution")
        print("  ✅ CT log queries")
        print("  ✅ Excel report generation")
        print("  ⚠️  Port scanning may not work (requires npcap)")
        print("  ⚠️  Some tools may be missing\n")
    
    def check_command(self, cmd):
        try:
            if self.platform == "windows":
                result = subprocess.run(["where", cmd], capture_output=True, text=True)
            else:
                result = subprocess.run(["which", cmd], capture_output=True, text=True)
            return result.returncode == 0
        except:
            return False
    
    def _pip_install(self, packages: list, label: str = None) -> bool:
        """
        Install one or more pip packages using the best available method.
        Tried in order:
          1. pipx          -- preferred for CLI tools; sandboxed, no env pollution
          2. pip --break-system-packages  -- needed on macOS Homebrew Python / Debian bookworm+
          3. pip --user    -- fallback for older setups without PEP 668 enforcement
        Returns True if installation succeeded, False otherwise.
        """
        label = label or ", ".join(packages)

        # Strategy 1: pipx
        try:
            r = subprocess.run([sys.executable, "-m", "pipx", "--version"],
                               capture_output=True, text=True)
            if r.returncode == 0:
                all_ok = True
                for pkg in packages:
                    r2 = subprocess.run(
                        [sys.executable, "-m", "pipx", "install", pkg, "--quiet"],
                        capture_output=True, text=True
                    )
                    if r2.returncode not in (0, 1):   # 1 = already installed, that is fine
                        all_ok = False
                if all_ok:
                    print(f"  ✅ {label} installed via pipx")
                    return True
        except Exception:
            pass

        # Strategy 2: pip --break-system-packages  (macOS Homebrew Python / PEP 668)
        try:
            r = subprocess.run(
                [sys.executable, "-m", "pip", "install", "--quiet",
                 "--break-system-packages"] + packages,
                capture_output=True, text=True
            )
            if r.returncode == 0:
                print(f"  ✅ {label} installed via pip (--break-system-packages)")
                return True
        except Exception:
            pass

        # Strategy 3: pip --user (older setups without PEP 668 enforcement)
        try:
            r = subprocess.run(
                [sys.executable, "-m", "pip", "install", "--quiet", "--user"] + packages,
                capture_output=True, text=True
            )
            if r.returncode == 0:
                print(f"  ✅ {label} installed via pip --user")
                return True
        except Exception:
            pass

        # All strategies failed -- print platform-specific hint
        print(f"  ⚠️  Could not auto-install {label}.")
        if self.platform == "darwin":
            print(f"     Try manually:  brew install pipx && pipx install {packages[0]}")
            print(f"     Or:            pip install --break-system-packages {' '.join(packages)}")
        else:
            print(f"     Try manually:  pip install --user {' '.join(packages)}")
            print(f"     Or:            pipx install {packages[0]}")
        return False

    def ensure_python_deps(self):
        """Install required Python library dependencies (imported in code, not CLI tools)."""
        packages = {
            "openpyxl":   "openpyxl",
            "tldextract": "tldextract",
            "ipwhois":    "ipwhois",
        }
        missing = []
        for module, package in packages.items():
            try:
                __import__(module)
            except ImportError:
                missing.append(package)
        if missing:
            print(f"ð¦ Installing Python library dependencies: {', '.join(missing)}...")
            self._pip_install(missing, label=", ".join(missing))
            print("   (if this failed, run: pip install " + " ".join(missing) + " --break-system-packages)")

    def install_go(self):
        print("ð¦ Go not found. Installing Go...")
        
        if self.platform == "darwin":
            print("   Detected macOS")
            if self.check_command("brew"):
                print("   Installing via Homebrew...")
                subprocess.run(["brew", "install", "go"], check=False)
            else:
                print("   ⚠️  Homebrew not found")
                print("   Install Homebrew: https://brew.sh")
                print("   Or install Go manually: https://go.dev/dl/")
                
                response = input("\n   Continue without Go tools? [y/N]: ")
                if response.lower() not in ['y', 'yes']:
                    sys.exit(1)
                return
                
        elif self.platform in ["linux", "wsl2"]:
            print(f"   Detected {self.platform.upper()}")
            
            try:
                import urllib.request
                with urllib.request.urlopen("https://go.dev/VERSION?m=text", timeout=10) as resp:
                    go_version = resp.read().decode().strip().splitlines()[0].lstrip("go")
            except Exception:
                go_version = "1.21.6"

            arch = platform.machine()
            
            if arch == "x86_64":
                go_arch = "amd64"
            elif arch == "aarch64" or arch == "arm64":
                go_arch = "arm64"
            else:
                print(f"   ⚠️  Unsupported architecture: {arch}")
                return
            
            go_tarball = f"go{go_version}.linux-{go_arch}.tar.gz"
            go_url = f"https://go.dev/dl/{go_tarball}"
            
            print(f"   Downloading Go {go_version}...")
            result = subprocess.run(["curl", "-L", "-o", f"/tmp/{go_tarball}", go_url], 
                                  capture_output=True)
            
            if result.returncode != 0:
                print("   ⚠️  Failed to download Go")
                return
            
            print("   Installing Go to ~/go...")
            subprocess.run(["tar", "-C", str(self.home), "-xzf", 
                          f"/tmp/{go_tarball}"], check=False)
            
            shell_rc = self.home / ".bashrc"
            if (self.home / ".zshrc").exists():
                shell_rc = self.home / ".zshrc"
            
            go_path_export = f'\nexport PATH="$HOME/go/bin:$PATH"\n'
            
            try:
                with open(shell_rc, "a") as f:
                    existing_content = open(shell_rc).read() if shell_rc.exists() else ""
                    if go_path_export.strip() not in existing_content:
                        f.write(go_path_export)
            except:
                pass
            
            os.environ["PATH"] = f"{self.home}/go/bin:{os.environ['PATH']}"
            
            print("✅ Go installed successfully")
    
    def ensure_go_tool(self, tool_name, go_package):
        if not self.check_command(tool_name):
            tool_path = self.go_path / tool_name
            if not tool_path.exists():
                print(f"ð¦ Installing {tool_name}...")
                try:
                    result = subprocess.run(["go", "install", "-v", go_package], 
                                          capture_output=True,
                                          text=True,
                                          timeout=120)
                    if result.returncode == 0:
                        print(f"✅ {tool_name} installed")
                    else:
                        print(f"⚠️  Failed to install {tool_name}")
                except subprocess.TimeoutExpired:
                    print(f"⚠️  Timeout installing {tool_name}")
                except Exception as e:
                    print(f"⚠️  Error installing {tool_name}: {e}")
    
    def ensure_pd_httpx(self):
        """
        Install ProjectDiscovery httpx and store its resolved path.
        Problem: Python's 'httpx' pip package also installs an 'httpx' binary,
        so we can't just check 'which httpx' — we might get the wrong one.
        Resolution order:
          1. ~/go/bin/httpx       (go install always lands here)
          2. which httpx          (only trusted if it identifies as PD httpx)
        The resolved path is stored in self._pd_httpx_bin for use in phases.
        """
        go_bin = self.go_path / "httpx"

        # Already installed in go/bin
        if go_bin.exists():
            self._pd_httpx_bin = str(go_bin)
            return

        # Check if the httpx on PATH is the PD version
        try:
            r = subprocess.run(["httpx", "-version"], capture_output=True, text=True, timeout=5)
            if "projectdiscovery" in (r.stdout + r.stderr).lower():
                self._pd_httpx_bin = "httpx"
                return
        except Exception:
            pass

        # Not installed — install it
        print("ð¦ Installing projectdiscovery/httpx...")
        try:
            result = subprocess.run(
                ["go", "install", "-v",
                 "github.com/projectdiscovery/httpx/cmd/httpx@latest"],
                capture_output=True, text=True, timeout=120
            )
            if result.returncode == 0:
                print("✅ httpx (projectdiscovery) installed")
                self._pd_httpx_bin = str(go_bin)
            else:
                print("⚠️  Failed to install httpx (projectdiscovery)")
                self._pd_httpx_bin = None
        except Exception as e:
            print(f"⚠️  Error installing httpx: {e}")
            self._pd_httpx_bin = None

    def ensure_python_tool(self, tool_name: str, pip_package: str):
        """
        Install a Python CLI tool (one that lands on PATH after pip install).
        Checks for the binary first; if missing, delegates to _pip_install()
        which tries pipx → pip --break-system-packages → pip --user in order.
        """
        # Check both PATH and ~/.local/bin (common pip --user location)
        local_bin = self.home / ".local" / "bin" / tool_name
        if self.check_command(tool_name) or local_bin.exists():
            return
        print(f"ð¦ Installing {pip_package}...")
        self._pip_install([pip_package], label=pip_package)

    def ensure_365doms(self):
        """Fetch 365doms.py from GitHub gist if not already present"""
        script_path = self.home / "tools" / "365doms.py"
        script_path.parent.mkdir(parents=True, exist_ok=True)
        if script_path.exists():
            return
        url = ("https://gist.githubusercontent.com/nullenc0de/2981930386c439af08c031622440bc2e"
               "/raw/47c8a04e058ccd7c47f9b59b8078694af64fd520/365doms.py")
        print(f"ð¦ Fetching 365doms.py from gist...")
        try:
            import urllib.request
            urllib.request.urlretrieve(url, str(script_path))
            script_path.chmod(0o755)
            print(f"✅ 365doms.py saved to {script_path}")
        except Exception as e:
            print(f"⚠️  Could not fetch 365doms.py: {e}")

    def check_system_tool(self, tool_name):
        if not self.check_command(tool_name):
            print(f"⚠️  {tool_name} not found")
            
            if self.platform == "darwin":
                print(f"   Install with: brew install {tool_name}")
            elif self.platform in ["linux", "wsl2"]:
                if tool_name == "dig":
                    print(f"   Install with: sudo apt install dnsutils  (Debian/Ubuntu)")
                    print(f"              or: sudo yum install bind-utils  (RHEL/CentOS)")
                else:
                    print(f"   Install with: sudo apt install {tool_name}  (Debian/Ubuntu)")
                    print(f"              or: sudo yum install {tool_name}  (RHEL/CentOS)")


class ReconFramework:
    def __init__(self, targets: List[str], output_dir: str = "recon_output",
                 github_token: str = None, shodan_token: str = None, pdcp_token: str = None):
        self.targets = targets
        self.output_dir = output_dir
        self.github_token = github_token
        self.shodan_token = shodan_token
        self.pdcp_token = pdcp_token
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.results = {
            "subdomains": set(),
            "github_subdomains": set(),      # github-subdomains results
            "gau_subdomains": set(),          # gau URL-extracted subdomains
            "shodan_subdomains": set(),       # shosubgo / shodan results
            "sublist3r_subdomains": set(),    # sublist3r results
            "asnmap_cidrs": set(),            # asnmap CIDR ranges (apex enrichment)
            "httpx_results": [],              # httpx live host probes
            "ip_addresses": set(),
            "open_ports": {},
            "apex_domains": set(),
            "ct_logs": set(),
            "services": {},
            "web_crawl": set(),
            "reverse_dns_domains": set(),
            "ip_enrichment": {},
            "cdn_results": {},
            "tools_called": {},
            "cidr_ranges": [],
            "cidr_expanded_ips": set(),
            "phases_executed": {
                "subdomains": False,
                "github_subdomains": False,
                "gau_subdomains": False,
                "shodan_subdomains": False,
                "sublist3r_subdomains": False,
                "asnmap": False,
                "httpx_probe": False,
                "bruteforce": False,
                "dns": False,
                "ports": False,
                "ct_logs": False,
                "apex": False,
                "services": False,
                "web_crawl": False,
                "reverse_dns": False,
                "ip_enrichment": False,
                "cdn_check": False,
                "cidr_expansion": False
            }
        }
        
        self._pd_httpx_bin = None   # resolved in ensure_pd_httpx()
        os.makedirs(self.output_dir, exist_ok=True)
    
    def log(self, message: str, level: str = "INFO"):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{timestamp}] [{level}] {message}")
    
    def save_results(self, phase: str, data: any):
        output_file = os.path.join(self.output_dir, f"{phase}_{self.timestamp}.txt")
        with open(output_file, 'w') as f:
            if isinstance(data, (set, list)):
                f.write('\n'.join(str(item) for item in data))
            elif isinstance(data, dict):
                f.write(json.dumps(data, indent=2))
            else:
                f.write(str(data))
        self.log(f"Results saved to {output_file}")
        return output_file
    
    def run_command(self, cmd: List[str], shell: bool = False, timeout: int = 300) -> tuple:
        try:
            if shell:
                result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=timeout)
            else:
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
            return result.returncode, result.stdout, result.stderr
        except subprocess.TimeoutExpired:
            self.log(f"Command timed out: {' '.join(cmd) if isinstance(cmd, list) else cmd}", "ERROR")
            return -1, "", "Timeout"
        except Exception as e:
            self.log(f"Command failed: {str(e)}", "ERROR")
            return -1, "", str(e)
    
    def check_tool(self, tool: str) -> bool:
        try:
            if platform.system().lower() == "windows":
                result = subprocess.run(["where", tool], capture_output=True, text=True)
            else:
                result = subprocess.run(["which", tool], capture_output=True, text=True)
            return result.returncode == 0
        except:
            return False
    
    def categorize_targets(self) -> Dict[str, List[str]]:
        ip_pattern = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')
        cidr_pattern = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}/\d{1,2}$')

        categorized = {"ips": [], "cidrs": [], "domains": []}

        for target in self.targets:
            target = target.strip()
            if ip_pattern.match(target):
                categorized["ips"].append(target)
            elif cidr_pattern.match(target):
                categorized["cidrs"].append(target)
            else:
                categorized["domains"].append(target)

        return categorized
    
    def expand_cidr_targets(self, cidrs: List[str]) -> Set[str]:
        self.results["phases_executed"]["cidr_expansion"] = True
        expanded_ips = set()
        LARGE_RANGE_THRESHOLD = 1024

        for cidr in cidrs:
            try:
                import ipaddress
                network = ipaddress.ip_network(cidr, strict=False)

                if network.prefixlen == 32:
                    expanded_ips.add(str(network.network_address))
                    self.log(f"CIDR {cidr} → single host (treated as /32)")
                    continue

                host_count = sum(1 for _ in network.hosts())

                if host_count > LARGE_RANGE_THRESHOLD:
                    self.log(f"⚠️  Large IP range detected: {cidr} ({host_count} hosts)", "WARNING")
                    response = input(f"   Continue expanding {cidr}? [y/N]: ")
                    if response.lower() not in ['y', 'yes']:
                        self.log(f"Skipping {cidr}", "WARNING")
                        continue

                hosts = [str(ip) for ip in network.hosts()]
                expanded_ips.update(hosts)
                self.log(f"Expanding CIDR: {cidr} → {len(hosts)} hosts")

            except ValueError as e:
                self.log(f"Invalid CIDR notation '{cidr}': {e} - skipping", "ERROR")

        self.results["cidr_ranges"] = cidrs
        self.results["cidr_expanded_ips"] = expanded_ips
        self.log(f"Total IPs from CIDR expansion: {len(expanded_ips)}")
        return expanded_ips
    
    @staticmethod
    def _sanitize(value: str) -> str:
        """Strip ANSI escape codes and openpyxl-illegal control characters.
        Tools like sublist3r leak colour codes even with --no-color/-n."""
        import re as _re
        # ANSI escape sequences e.g. \x1b[31m
        # bracket-terminated first: sublist3r timestamps like \x1b[15:09:48]
        value = _re.sub(r'\x1b\[[^\]]*\]\s*', '', value)
        # letter-terminated: standard ANSI colour/cursor sequences
        value = _re.sub(r'\x1b\[[0-9;:]*[a-zA-Z]', '', value)
        # bare ESC + one char (catch anything left)
        value = _re.sub(r'\x1b.', '', value)
        # Any remaining C0/C1 control chars except tab/newline
        value = _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', value)
        return value.strip()

    @staticmethod
    def _is_valid_hostname(value: str) -> bool:
        """
        Return True only if value looks like a real hostname/domain.
        Rejects URLs, GitHub links, search queries, timestamps, log lines,
        bare TLDs, wildcards, and anything with illegal hostname characters.
        """
        import re as _re
        if not value:
            return False
        # Reject anything that looks like a URL (has a scheme)
        if '://' in value:
            return False
        # Reject percent-encoded strings (URL query params)
        if '%' in value:
            return False
        # Reject lines starting with [ (tool log lines like [16:18:27])
        if value.startswith('['):
            return False
        # Reject wildcards
        if value.startswith('*'):
            return False
        # Must contain at least one dot
        if '.' not in value:
            return False
        # Must not contain spaces, slashes, colons, @, query chars
        if _re.search(r'[ /\\:@?=&,<>()\[\]{}|\'"]', value):
            return False
        # Each label must be non-empty, ≤63 chars, only alnum/hyphen/underscore
        # (underscore allowed for _dmarc, _domainkey etc)
        labels = value.split('.')
        if len(labels) < 2:
            return False
        for label in labels:
            if not label:
                return False
            if len(label) > 63:
                return False
            if not _re.match(r'^[a-zA-Z0-9_\-]+$', label):
                return False
        # TLD must be at least 2 chars and only letters
        if not _re.match(r'^[a-zA-Z]{2,}$', labels[-1]):
            return False
        return True

    def deduplicate_results(self):
        def _clean_set(key, lower=True):
            cleaned = set()
            for d in self.results.get(key, set()):
                if not d:
                    continue
                d = self._sanitize(str(d))
                if not d:
                    continue
                if lower:
                    d = d.lower()
                if not self._is_valid_hostname(d):
                    continue
                cleaned.add(d)
            self.results[key] = cleaned

        _clean_set("subdomains")
        _clean_set("github_subdomains")
        _clean_set("gau_subdomains")
        _clean_set("shodan_subdomains")
        _clean_set("sublist3r_subdomains")
        _clean_set("ct_logs")
        _clean_set("apex_domains")
        _clean_set("web_crawl")
        _clean_set("reverse_dns_domains")
        self.results["ip_addresses"] = {
            self._sanitize(ip) for ip in self.results.get("ip_addresses", set())
            if ip and self._sanitize(str(ip))
        }
        self.log("Results deduplicated")

    def phase_subdomain_enum(self) -> Set[str]:
        """
        Phase 1: Unified Subdomain Enumeration
        All tools run in parallel; results merged into one deduplicated list.
          - subfinder          passive, broadest source coverage
          - amass              additional passive sources
          - assetfinder        --subs-only flag (apex discovery handled separately in phase_apex)
          - gau                URL archive mining → hostname extraction
          - github-subdomains  GitHub code search (requires --github-token)
        Everything lands in self.results["subdomains"].
        Per-tool subsets stored separately for audit / Excel reporting.
        """
        self.log("Phase 1: Starting subdomain enumeration (all tools in parallel)")
        self.results["phases_executed"]["subdomains"] = True
        self.results["phases_executed"]["github_subdomains"] = True
        self.results["phases_executed"]["gau_subdomains"] = True

        domain_targets = [
            t for t in self.targets
            if not re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}(/\d+)?$', t)
        ]

        if not domain_targets:
            self.log("No domain targets to enumerate subdomains for", "WARNING")
            return set()

        # Shared temp file used by batch tools
        temp_domains = os.path.join(self.output_dir, f"temp_targets_{self.timestamp}.txt")
        with open(temp_domains, 'w') as f:
            f.write('\n'.join(domain_targets))

        all_subdomains: Set[str]      = set()
        github_subdomains: Set[str]   = set()
        gau_subdomains: Set[str]      = set()
        shodan_subdomains: Set[str]   = set()
        sublist3r_subdomains: Set[str]= set()

        # ── helpers ───────────────────────────────────────────────────────────

        def _clean(lines) -> Set[str]:
            """Lowercase, strip, drop empties."""
            return {l.strip().lower() for l in lines if l.strip()}

        def run_subfinder() -> Set[str]:
            if not self.check_tool("subfinder"):
                self.log("[subfinder] not found — skipping", "WARNING")
                return set()
            self.log(f"[subfinder] Running on {len(domain_targets)} domain(s)...")
            cmd = ["subfinder", "-dL", temp_domains, "-silent", "-all"]
            self._record_tool("subfinder", cmd)
            rc, stdout, _ = self.run_command(cmd, timeout=600)
            raw = _clean(stdout.splitlines()) if rc == 0 and stdout.strip() else set()
            found = {d for d in raw if self._is_valid_hostname(d)}
            self.log(f"[subfinder] Found {len(found)} subdomains")
            return found

        def run_amass() -> Set[str]:
            if not self.check_tool("amass"):
                return set()
            self.log(f"[amass] Running passive enum on {len(domain_targets)} domain(s)...")
            cmd = ["amass", "enum", "-passive", "-df", temp_domains]
            self._record_tool("amass", cmd)
            rc, stdout, _ = self.run_command(cmd, timeout=600)
            raw = _clean(stdout.splitlines()) if rc == 0 and stdout.strip() else set()
            found = {d for d in raw if self._is_valid_hostname(d)}
            self.log(f"[amass] Found {len(found)} subdomains")
            return found

        def run_assetfinder() -> Set[str]:
            """
            Runs assetfinder --subs-only for subdomain discovery.
            The no-flag run (for apex) is handled in phase_apex_discovery().
            """
            if not self.check_tool("assetfinder"):
                return set()
            self.log(f"[assetfinder] Running --subs-only on {len(domain_targets)} domain(s)...")
            found = set()
            for domain in domain_targets:
                cmd = ["assetfinder", "--subs-only", domain]
                self._record_tool("assetfinder (subs-only)", cmd)
                rc, stdout, _ = self.run_command(cmd, timeout=120)
                if rc == 0 and stdout.strip():
                    found.update(d for d in _clean(stdout.splitlines()) if self._is_valid_hostname(d))
            self.log(f"[assetfinder] Found {len(found)} subdomains")
            return found

        def run_gau() -> Set[str]:
            """
            gau mines URL archives (Wayback, CommonCrawl, etc.) for each domain,
            then extracts unique hostnames from the URLs.
            Equivalent bash:
              echo $domain | gau --subs | cut -d/ -f3 | cut -d: -f1 | sort -u
            """
            if not self.check_tool("gau"):
                return set()
            self.log(f"[gau] Mining URL archives for {len(domain_targets)} domain(s)...")
            found = set()
            for domain in domain_targets:
                cmd = ["gau", "--subs", domain]
                self._record_tool("gau", cmd)
                rc, stdout, _ = self.run_command(cmd, timeout=300)
                if rc == 0 and stdout.strip():
                    for url in stdout.splitlines():
                        url = url.strip()
                        if not url:
                            continue
                        # Extract hostname: strip scheme, take host part, drop port
                        try:
                            # Fast manual parse — avoids urllib import inside thread
                            if "://" in url:
                                host = url.split("://", 1)[1].split("/")[0].split(":")[0]
                            else:
                                host = url.split("/")[0].split(":")[0]
                            host = host.lower().strip()
                            if host and domain in host and self._is_valid_hostname(host):
                                found.add(host)
                        except Exception:
                            continue
            gau_subdomains.update(found)   # track separately for reporting
            self.log(f"[gau] Found {len(found)} unique hostnames from URL archives")
            return found

        def run_github_subdomains() -> Set[str]:
            token = self.github_token
            if not token:
                self.log("[github-subdomains] No token — skipping (use --github-token)", "WARNING")
                return set()
            if not self.check_tool("github-subdomains"):
                self.log("[github-subdomains] Not installed — skipping", "WARNING")
                self.log("    Install: go install github.com/gwen001/github-subdomains@latest", "WARNING")
                return set()
            found = set()
            for domain in domain_targets:
                self.log(f"[github-subdomains] Searching GitHub for {domain}...")
                out_file = os.path.join(self.output_dir, f"github_subs_{domain}_{self.timestamp}.txt")
                cmd = ["github-subdomains", "-d", domain, "-t", token, "-o", out_file]
                cmd_display = ["github-subdomains", "-d", domain, "-t", "<REDACTED>", "-o", out_file]
                self._record_tool("github-subdomains", cmd_display)
                rc, stdout, _ = self.run_command(cmd, timeout=120)
                if os.path.exists(out_file):
                    with open(out_file) as fh:
                        for line in fh:
                            line = self._sanitize(line).lower()
                            if self._is_valid_hostname(line):
                                found.add(line)
                if stdout.strip():
                    for line in stdout.splitlines():
                        line = self._sanitize(line).lower()
                        if self._is_valid_hostname(line):
                            found.add(line)
            github_subdomains.update(found)   # track separately for reporting
            self.log(f"[github-subdomains] Found {len(found)} subdomains")
            return found

        def run_sublist3r() -> Set[str]:
            """
            Sublist3r: installed via pip as 'sublist3r'.
            CLI: sublist3r -d domain -o output.txt -n (no color)
            """
            if not self.check_tool("sublist3r"):
                self.log("[sublist3r] Not installed — skipping", "WARNING")
                self.log("    Install: pip install sublist3r", "WARNING")
                return set()
            found = set()
            for domain in domain_targets:
                self.log(f"[sublist3r] Enumerating {domain}...")
                out_file = os.path.join(self.output_dir, f"sublist3r_{domain}_{self.timestamp}.txt")
                cmd = ["sublist3r", "-d", domain, "-o", out_file, "-n"]
                self._record_tool("sublist3r", cmd)
                rc, stdout, _ = self.run_command(cmd, timeout=300)
                # Read output file (sublist3r writes full hostnames there)
                if os.path.exists(out_file):
                    with open(out_file) as fh:
                        for line in fh:
                            line = line.strip().lower()
                            if line and "." in line and not line.startswith("#"):
                                found.add(line)
                # Also parse stdout as fallback
                if stdout.strip() and not found:
                    for line in stdout.splitlines():
                        line = line.strip().lower()
                        # Strip ANSI escape codes that slip through -n
                        line = re.sub(r"\[[0-9;]*m", "", line)
                        if line and "." in line and domain in line:
                            found.add(line)
            sublist3r_subdomains.update(found)
            self.log(f"[sublist3r] Found {len(found)} subdomains")
            return found

        def run_shosubgo() -> Set[str]:
            """
            shosubgo: Shodan-backed subdomain discovery.
            CLI: shosubgo -d domain -s APIKEY
            Requires --shodan-token.
            """
            token = self.shodan_token
            if not token:
                self.log("[shosubgo] No Shodan API key — skipping (use --shodan-token)", "WARNING")
                return set()
            if not self.check_tool("shosubgo"):
                self.log("[shosubgo] Not installed — skipping", "WARNING")
                self.log("    Install: go install github.com/incogbyte/shosubgo@latest", "WARNING")
                return set()
            found = set()
            for domain in domain_targets:
                self.log(f"[shosubgo] Querying Shodan for {domain}...")
                out_file = os.path.join(self.output_dir, f"shosubgo_{domain}_{self.timestamp}.txt")
                cmd = ["shosubgo", "-d", domain, "-s", token, "-o", out_file]
                cmd_display = ["shosubgo", "-d", domain, "-s", "<REDACTED>", "-o", out_file]
                self._record_tool("shosubgo", cmd_display)
                rc, stdout, _ = self.run_command(cmd, timeout=120)
                if os.path.exists(out_file):
                    with open(out_file) as fh:
                        for line in fh:
                            line = line.strip().lower()
                            if line and "." in line and not line.startswith("#"):
                                found.add(line)
                if stdout.strip():
                    for line in stdout.splitlines():
                        line = line.strip().lower()
                        if line and "." in line and domain in line:
                            found.add(line)
            shodan_subdomains.update(found)
            self.log(f"[shosubgo] Found {len(found)} subdomains")
            return found

        # ── run all tools concurrently ────────────────────────────────────────
        workers = {
            "subfinder":         run_subfinder,
            "amass":             run_amass,
            "assetfinder":       run_assetfinder,
            "gau":               run_gau,
            "sublist3r":         run_sublist3r,
            "shosubgo":          run_shosubgo,
            "github-subdomains": run_github_subdomains,
        }

        with ThreadPoolExecutor(max_workers=len(workers)) as executor:
            futures = {executor.submit(fn): name for name, fn in workers.items()}
            for future in as_completed(futures):
                tool_name = futures[future]
                try:
                    all_subdomains.update(future.result())
                except Exception as e:
                    self.log(f"[{tool_name}] Error: {e}", "WARNING")

        if os.path.exists(temp_domains):
            os.remove(temp_domains)

        all_subdomains = {s for s in all_subdomains if s}
        self.results["subdomains"]         = all_subdomains
        self.results["github_subdomains"]  = github_subdomains
        self.results["gau_subdomains"]     = gau_subdomains
        self.results["shodan_subdomains"]  = shodan_subdomains
        self.results["sublist3r_subdomains"]= sublist3r_subdomains

        passive = all_subdomains - github_subdomains - gau_subdomains - shodan_subdomains - sublist3r_subdomains
        self.log(f"Subdomain enumeration complete: {len(all_subdomains)} unique subdomains")
        self.log(f"  • subfinder/amass/assetfinder: {len(passive)}")
        self.log(f"  • gau URL archives:            {len(gau_subdomains)}")
        self.log(f"  • sublist3r:                   {len(sublist3r_subdomains)}")
        self.log(f"  • shosubgo (Shodan):           {len(shodan_subdomains)}")
        self.log(f"  • github-subdomains:           {len(github_subdomains)}")

        # Save only valid hostnames to the subdomains .txt file
        clean_for_save = {d for d in all_subdomains if self._is_valid_hostname(d)}
        self.save_results("subdomains", clean_for_save)
        for key, data in [
            ("github_subdomains",   github_subdomains),
            ("gau_subdomains",      gau_subdomains),
            ("shodan_subdomains",   shodan_subdomains),
            ("sublist3r_subdomains",sublist3r_subdomains),
        ]:
            if data:
                self.save_results(key, data)

        return all_subdomains

    def phase_httpx_probe(self, domains: Set[str] = None) -> List[Dict]:
        """
        HTTP probing with ProjectDiscovery httpx on common web ports.
        Probes: 80, 443, 8080, 8443, 8888, 9443

        Binary resolution (avoids conflict with Python httpx package):
          1. ~/go/bin/httpx  -- go install always puts it here
          2. httpx on PATH   -- only if it reports as projectdiscovery
        Input: uses saved subdomains_*.txt file first (already cleaned),
               falls back to results dict if file not found.
        """
        self.log("Phase httpx: Probing web ports on discovered domains")
        self.results["phases_executed"]["httpx_probe"] = True

        # ── Resolve binary ────────────────────────────────────────────────────
        httpx_bin = None
        go_httpx = Path.home() / "go" / "bin" / "httpx"
        if go_httpx.exists():
            httpx_bin = str(go_httpx)
        else:
            try:
                r = subprocess.run(["httpx", "-version"],
                                   capture_output=True, text=True, timeout=5)
                if "projectdiscovery" in (r.stdout + r.stderr).lower():
                    httpx_bin = "httpx"
            except Exception:
                pass

        if not httpx_bin:
            self.log("⚠️  ProjectDiscovery httpx not found — skipping web probe", "WARNING")
            self.log("    Install: go install github.com/projectdiscovery/httpx/cmd/httpx@latest", "WARNING")
            return []

        # ── Build domain list ─────────────────────────────────────────────────
        # Primary: use the saved subdomains .txt file — already cleaned/validated.
        # Fall back to results dict if no file found.
        domain_set: Set[str] = set()

        saved_subs_file = None
        for fname in sorted(os.listdir(self.output_dir)):
            if fname.startswith("subdomains_") and fname.endswith(".txt"):
                saved_subs_file = os.path.join(self.output_dir, fname)
                # keep iterating to get the most recent (sorted alpha = sorted by timestamp)

        if saved_subs_file:
            self.log(f"[httpx] Reading domains from {saved_subs_file}")
            with open(saved_subs_file) as f:
                for line in f:
                    h = line.strip().lower()
                    if h and self._is_valid_hostname(h):
                        domain_set.add(h)

        # Also add original input targets (domains only)
        categorized = self.categorize_targets()
        for d in categorized["domains"]:
            h = d.strip().lower()
            if h and self._is_valid_hostname(h):
                domain_set.add(h)

        # If we have an explicit domains argument, merge it in too
        if domains:
            for d in domains:
                h = self._sanitize(str(d)).lower()
                if h and self._is_valid_hostname(h):
                    domain_set.add(h)

        # Last resort — pull directly from results dict
        if not domain_set:
            for key in ("subdomains", "ct_logs", "reverse_dns_domains",
                        "github_subdomains", "gau_subdomains",
                        "sublist3r_subdomains", "shodan_subdomains"):
                for d in self.results.get(key, set()):
                    h = self._sanitize(str(d)).lower()
                    if h and self._is_valid_hostname(h):
                        domain_set.add(h)

        if not domain_set:
            self.log("[httpx] No valid domains to probe — skipping", "WARNING")
            return []

        # ── Write input file ──────────────────────────────────────────────────
        perm_input = os.path.join(self.output_dir, f"httpx_domains_{self.timestamp}.txt")
        with open(perm_input, "w") as f:
            f.write("\n".join(sorted(domain_set)))
        self.log(f"[httpx] {len(domain_set)} domains → {perm_input}")

        httpx_output = os.path.join(self.output_dir, f"httpx_probe_{self.timestamp}.json")
        WEB_PORTS = "80,443,8080,8443,8888,9443"

        cmd = [
            httpx_bin,
            "-l",       perm_input,
            "-p",       WEB_PORTS,
            "-title",
            "-sc",
            "-server",
            "-td",
            "-ip",
            "-cdn",
            "-nc",
            "-silent",
            "-json",
            "-o",       httpx_output,
            "-fr",
            "-timeout", "10",
            "-t",       "50",
            "-retries", "1",
        ]
        cmd_display = ["httpx"] + cmd[1:]  # use plain "httpx" in audit log
        self._record_tool("httpx", cmd_display)
        self.log(f"[httpx] Running on ports {WEB_PORTS}...")
        returncode, stdout, stderr = self.run_command(cmd, timeout=600)

        if returncode != 0:
            self.log(f"[httpx] Non-zero exit ({returncode}) — stderr: {stderr[:200]}", "WARNING")

        results = []
        if os.path.exists(httpx_output):
            with open(httpx_output) as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        entry = json.loads(line)
                        results.append({
                            "url":         entry.get("url", ""),
                            "status_code": entry.get("status_code", ""),
                            "title":       self._sanitize(entry.get("title", "") or ""),
                            "server":      entry.get("webserver", "") or entry.get("server", ""),
                            "tech":        ", ".join(entry.get("tech", []) or []),
                            "ip":          entry.get("host", "") or entry.get("ip", ""),
                            "cdn":         entry.get("cdn_name", "") or "",
                        })
                    except json.JSONDecodeError:
                        pass
        else:
            self.log(f"[httpx] Output file not found: {httpx_output}", "WARNING")

        live_count = len(results)
        self.log(f"[httpx] {live_count} live web endpoints found")
        for r in results[:10]:
            status = r["status_code"]
            title  = f" — {r['title']}" if r["title"] else ""
            server = f" [{r['server']}]" if r["server"] else ""
            self.log(f"  {r['url']} [{status}]{server}{title}")
        if live_count > 10:
            self.log(f"  ... and {live_count - 10} more")

        self.results["httpx_results"] = results
        self.save_results("httpx_probe", results)
        return results

    def phase_reverse_dns(self, ips: Set[str] = None) -> Set[str]:
        """Phase 1b: Reverse DNS Lookup to discover domains from IPs"""
        self.log("Phase 1b: Performing reverse DNS lookups")
        self.results["phases_executed"]["reverse_dns"] = True
        
        discovered_domains = set()
        
        if ips is None:
            ips = self.results["ip_addresses"]
        
        if not ips:
            self.log("No IPs to perform reverse DNS on", "WARNING")
            return discovered_domains
        
        if self.check_tool("dnsx"):
            self.log(f"Performing reverse DNS on {len(ips)} IPs (batch mode)")
            
            temp_ips = os.path.join(self.output_dir, f"temp_reverse_ips_{self.timestamp}.txt")
            with open(temp_ips, 'w') as f:
                f.write('\n'.join(ips))
            
            cmd = ["dnsx", "-l", temp_ips, "-ptr", "-resp-only", "-silent"]
            self._record_tool("dnsx (reverse PTR)", cmd)
            returncode, stdout, _ = self.run_command(cmd, timeout=300)
            
            if returncode == 0 and stdout.strip():
                for line in stdout.strip().split('\n'):
                    if line and line.strip():
                        domain = line.strip().rstrip('.')
                        if not any(pattern in domain.lower() for pattern in [
                            '.in-addr.arpa', '.ip6.arpa', 'static', 'dynamic', 
                            'dhcp', 'pool', 'dsl', 'cable'
                        ]):
                            discovered_domains.add(domain)
            
            if os.path.exists(temp_ips):
                os.remove(temp_ips)
        
        elif self.check_tool("dig"):
            self.log(f"Performing reverse DNS on {len(ips)} IPs (individual lookups)")
            for ip in list(ips)[:50]:
                dig_cmd = ["dig", "+short", "-x", ip]
                self._record_tool("dig (reverse PTR fallback)", dig_cmd)
                returncode, stdout, _ = self.run_command(dig_cmd)
                if returncode == 0 and stdout.strip():
                    domain = stdout.strip().rstrip('.')
                    if not any(pattern in domain.lower() for pattern in [
                        '.in-addr.arpa', '.ip6.arpa', 'static', 'dynamic', 
                        'dhcp', 'pool', 'dsl', 'cable'
                    ]):
                        discovered_domains.add(domain)
        
        else:
            self.log("No reverse DNS tool available (dnsx or dig)", "WARNING")
        
        self.results["reverse_dns_domains"] = discovered_domains
        self.log(f"Discovered {len(discovered_domains)} domains via reverse DNS")
        if discovered_domains:
            for domain in list(discovered_domains)[:10]:
                self.log(f"  → {domain}")
            if len(discovered_domains) > 10:
                self.log(f"  ... and {len(discovered_domains) - 10} more")
        
        self.save_results("reverse_dns_domains", discovered_domains)
        return discovered_domains
    
    def phase_dns_resolution(self, domains: Set[str] = None) -> Set[str]:
        """Phase 2: DNS Resolution (FIXED - robust with fallback)"""
        self.log("Phase 2: Starting DNS resolution")
        self.results["phases_executed"]["dns"] = True
        
        if domains is None:
            domains = self.results["subdomains"] if self.results["subdomains"] else set(self.targets)
        
        ip_pattern = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')
        domains = {d for d in domains if not ip_pattern.match(d)}
        
        if not domains:
            self.log("No domains to resolve", "WARNING")
            return set()
        
        ip_addresses = set()
        
        if self.check_tool("dnsx"):
            temp_input = os.path.join(self.output_dir, f"temp_domains_{self.timestamp}.txt")
            with open(temp_input, 'w') as f:
                f.write('\n'.join(domains))
            
            self.log(f"Resolving {len(domains)} domains with dnsx")
            
            dnsx_cmd = ["dnsx", "-l", temp_input, "-a", "-resp-only", "-silent"]
            self._record_tool("dnsx (resolution)", dnsx_cmd)
            returncode, stdout, stderr = self.run_command(dnsx_cmd)
            
            if returncode == 0 and stdout.strip():
                for line in stdout.strip().split('\n'):
                    line = line.strip()
                    if line and re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', line):
                        ip_addresses.add(line)
            
            if not ip_addresses:
                self.log("Trying dnsx with different flags...", "INFO")
                dnsx_cmd2 = ["dnsx", "-l", temp_input, "-a", "-silent"]
                self._record_tool("dnsx (resolution)", dnsx_cmd2)
                returncode, stdout, stderr = self.run_command(dnsx_cmd2)
                
                if returncode == 0 and stdout.strip():
                    for line in stdout.strip().split('\n'):
                        if line:
                            ip_matches = re.findall(r'\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b', line)
                            if ip_matches:
                                ip_addresses.update(ip_matches)
            
            if os.path.exists(temp_input):
                os.remove(temp_input)
            
            if ip_addresses:
                self.log(f"dnsx resolved {len(ip_addresses)} IPs")
            else:
                self.log("dnsx returned no results, falling back to dig", "WARNING")
        
        if not ip_addresses and self.check_tool("dig"):
            self.log(f"Falling back to dig for {len(domains)} domains (slower)...")
            
            domains_list = list(domains)[:200]
            
            for idx, domain in enumerate(domains_list, 1):
                if idx % 20 == 0:
                    self.log(f"Resolving progress: {idx}/{len(domains_list)}")
                
                dig_cmd = ["dig", "+short", domain, "A"]
                self._record_tool("dig (resolution fallback)", dig_cmd)
                returncode, stdout, _ = self.run_command(dig_cmd)
                if returncode == 0 and stdout.strip():
                    for line in stdout.strip().split('\n'):
                        line = line.strip()
                        if re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', line):
                            ip_addresses.add(line)
            
            self.log(f"dig resolved {len(ip_addresses)} IPs")
        
        self.results["ip_addresses"] = self.results["ip_addresses"].union(ip_addresses)
        self.log(f"Resolved {len(ip_addresses)} new IP addresses (total: {len(self.results['ip_addresses'])})")
        
        if len(ip_addresses) > 0:
            sample_ips = list(ip_addresses)[:5]
            self.log(f"Sample IPs: {', '.join(sample_ips)}")
        else:
            self.log("WARNING: No IPs were resolved.", "WARNING")
        
        self.save_results("ip_addresses", self.results["ip_addresses"])
        return ip_addresses
    
    def phase_cdn_check(self, targets: Set[str] = None) -> Dict:
        """Identify CDN/WAF protected IPs and domains using cdncheck CLI"""
        self.log("Phase CDN: Checking for CDN/WAF protection")
        self.results["phases_executed"]["cdn_check"] = True

        cdn_results = {}

        if targets is None:
            targets = self.results["ip_addresses"].union(self.results["subdomains"])

        if not targets:
            self.log("No targets for CDN check", "WARNING")
            return cdn_results

        if not self.check_tool("cdncheck"):
            self.log("cdncheck not found. Install: go install github.com/projectdiscovery/cdncheck/cmd/cdncheck@latest", "WARNING")
            return cdn_results

        temp_input = os.path.join(self.output_dir, f"temp_cdn_{self.timestamp}.txt")
        with open(temp_input, 'w') as f:
            f.write('\n'.join(targets))

        cmd = ["cdncheck", "-i", temp_input, "-resp", "-silent"]
        self._record_tool("cdncheck", cmd)
        self.log(f"Checking {len(targets)} targets for CDN/WAF...")

        returncode, stdout, stderr = self.run_command(cmd, timeout=300)

        if returncode == 0 and stdout.strip():
            pattern = re.compile(r'^(\S+)\s+\[(\w+)\]\s+\[([^\]]+)\]')
            for line in stdout.strip().split('\n'):
                match = pattern.match(line.strip())
                if match:
                    target = match.group(1)
                    cdn_type = match.group(2)
                    provider = match.group(3)
                    cdn_results[target] = f"{provider} ({cdn_type})"

        if os.path.exists(temp_input):
            os.remove(temp_input)

        cdn_protected = {k: v for k, v in cdn_results.items() if v is not None}
        self.log(f"CDN/WAF check complete: {len(cdn_protected)} protected targets identified")
        for target, provider in list(cdn_protected.items())[:10]:
            self.log(f"  ð¡  {target} → {provider}")
        if len(cdn_protected) > 10:
            self.log(f"  ... and {len(cdn_protected) - 10} more")

        self.results["cdn_results"] = cdn_results
        self.save_results("cdn_results", cdn_results)
        return cdn_results

    def phase_ip_enrichment(self, ips: Set[str] = None) -> Dict:
        """Enrich IPs with ASN, org, country, network range via ipwhois RDAP"""
        self.log("Phase IP Enrichment: Querying RDAP/WHOIS for IP ownership info")
        self.results["phases_executed"]["ip_enrichment"] = True
        enrichment = {}

        if ips is None:
            ips = self.results["ip_addresses"]

        if not ips:
            self.log("No IPs to enrich", "WARNING")
            return enrichment

        try:
            from ipwhois import IPWhois
            self._record_tool("ipwhois/rdap", ["Python API: IPWhois(ip).lookup_rdap(depth=1)"])
            self.log(f"Enriching {len(ips)} IPs via RDAP (parallel, 5 workers)...")

            def enrich_single(ip):
                try:
                    result = IPWhois(ip).lookup_rdap(depth=1)
                    return ip, {
                        "asn": result.get("asn", "Unknown"),
                        "asn_description": result.get("asn_description", "Unknown"),
                        "organization": result.get("network", {}).get("name", "Unknown"),
                        "country": result.get("asn_country_code", "Unknown"),
                        "cidr": result.get("asn_cidr", "Unknown"),
                    }
                except Exception:
                    return ip, {"asn": "Unknown", "asn_description": "Unknown",
                                "organization": "Unknown", "country": "Unknown", "cidr": "Unknown"}

            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {executor.submit(enrich_single, ip): ip for ip in ips}
                for future in as_completed(futures):
                    ip, data = future.result()
                    enrichment[ip] = data

            self.log(f"IP enrichment complete for {len(enrichment)} IPs")

        except ImportError:
            self.log("ipwhois not installed. Run: pip install ipwhois", "WARNING")
        except Exception as e:
            self.log(f"IP enrichment error: {e}", "WARNING")

        self.results["ip_enrichment"] = enrichment
        self.save_results("ip_enrichment", enrichment)
        return enrichment

    def phase_port_scan(self, targets: Set[str] = None) -> Dict:
        """Phase 3: Port Scanning with naabu (two-pass: top 1000 + critical extras)"""
        self.log("Phase 3: Starting port scanning")
        self.results["phases_executed"]["ports"] = True
        
        if targets is None:
            targets = self.results["ip_addresses"] if self.results["ip_addresses"] else set(self.targets)
        
        open_ports = {}
        
        if not self.check_tool("naabu"):
            self.log("naabu not found. Install it with: go install -v github.com/projectdiscovery/naabu/v2/cmd/naabu@latest", "WARNING")
            return open_ports
        
        targets = {t.strip() for t in targets if t and t.strip()}
        
        if not targets:
            self.log("No valid targets for port scanning", "WARNING")
            return open_ports
        
        ip_pattern = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')
        ips = {t for t in targets if ip_pattern.match(t)}
        domains = {t for t in targets if not ip_pattern.match(t)}
        
        if domains:
            self.log(f"Detected {len(domains)} domain names, resolving to IPs first...")
            resolved_ips = self.phase_dns_resolution(domains)
            ips = ips.union(resolved_ips)
            self.log(f"Resolved {len(resolved_ips)} IPs from domains")
        
        if not ips:
            self.log("No valid IPs to scan after resolution", "WARNING")
            return open_ports
        
        temp_input = os.path.join(self.output_dir, f"temp_ips_{self.timestamp}.txt")
        with open(temp_input, 'w') as f:
            f.write('\n'.join(ips))
        
        self.log(f"Scanning {len(ips)} IPs with naabu")
        naabu_output = os.path.join(self.output_dir, f"naabu_{self.timestamp}.json")
        
        cdn_results = self.results.get("cdn_results", {})
        ip_pattern_check = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')
        cdn_ips = {ip for ip, provider in cdn_results.items() if provider is not None and ip_pattern_check.match(ip)}
        if cdn_ips:
            self.log(f"⚠️  Skipping {len(cdn_ips)} CDN-protected IPs from port scan:", "WARNING")
            for ip in cdn_ips:
                self.log(f"   ð¡  {ip} → {cdn_results[ip]}", "WARNING")
            ips = ips - cdn_ips

        if not ips:
            self.log("All IPs are CDN-protected, nothing to scan", "WARNING")
            return open_ports

        self.log("STEP 1: Scanning top 1000 common ports...")
        cmd_top = [
            "naabu",
            "-list", temp_input,
            "-top-ports", "1000",
            "-json",
            "-o", naabu_output,
            "-silent",
            "-rate", "5000",
            "-c", "50",
            "-retries", "1"
        ]
        
        self._record_tool("naabu (top 1000)", cmd_top)
        returncode, stdout, stderr = self.run_command(cmd_top, timeout=600)

        def _parse_naabu_port(raw) -> int:
            """
            Naabu JSON port field changed across versions:
              old: "port": 443          (int)
              new: "port": {"Port": 443, "Protocol": "tcp"}  (dict)
            Always returns an int, or None if unparseable.
            """
            if isinstance(raw, int):
                return raw
            if isinstance(raw, dict):
                # Try common key names across naabu versions
                for key in ("Port", "port", "number"):
                    if key in raw and isinstance(raw[key], int):
                        return raw[key]
            return None

        scanned_ports = set()
        if returncode == 0 and os.path.exists(naabu_output):
            with open(naabu_output, 'r') as f:
                for line in f:
                    if line.strip():
                        try:
                            data = json.loads(line.strip())
                            ip   = data.get('ip')
                            port = _parse_naabu_port(data.get('port'))

                            if ip and port:
                                if ip not in open_ports:
                                    open_ports[ip] = []
                                open_ports[ip].append(port)
                                scanned_ports.add(port)
                        except json.JSONDecodeError:
                            pass
        
        CRITICAL_EXTRA_PORTS = [
            2222, 2375, 2376, 2379, 2380, 4243, 4444,
            5000, 6379, 6443, 7001, 8009, 8888,
            9092, 9200, 9300, 10000, 10250, 10255,
            11211, 27017, 27018, 50000, 50070, 61616
        ]
        
        extra_ports = [p for p in CRITICAL_EXTRA_PORTS if p not in scanned_ports]
        
        if extra_ports:
            self.log(f"STEP 2: Scanning {len(extra_ports)} additional critical ports...")
            naabu_extra = os.path.join(self.output_dir, f"naabu_extra_{self.timestamp}.json")
            
            cmd_extra = [
                "naabu",
                "-list", temp_input,
                "-p", ",".join(map(str, extra_ports)),
                "-json",
                "-o", naabu_extra,
                "-silent",
                "-rate", "5000",
                "-c", "50",
                "-retries", "1"
            ]
            
            self._record_tool("naabu (critical ports)", cmd_extra)
            returncode, _, _ = self.run_command(cmd_extra, timeout=180)
            
            if returncode == 0 and os.path.exists(naabu_extra):
                with open(naabu_extra, 'r') as f:
                    for line in f:
                        if line.strip():
                            try:
                                data = json.loads(line.strip())
                                ip   = data.get('ip')
                                port = _parse_naabu_port(data.get('port'))

                                if ip and port:
                                    if ip not in open_ports:
                                        open_ports[ip] = []
                                    open_ports[ip].append(port)
                            except json.JSONDecodeError:
                                pass
        
        if os.path.exists(temp_input):
            os.remove(temp_input)
        
        for ip in open_ports:
            open_ports[ip] = sorted(open_ports[ip])
        
        self.results["open_ports"] = open_ports
        total_ports = sum(len(ports) for ports in open_ports.values())
        self.log(f"Found {total_ports} open ports across {len(open_ports)} hosts")
        self.save_results("open_ports", open_ports)
        return open_ports
    
    def phase_ct_logs(self) -> Set[str]:
        """Phase 4: Certificate Transparency Log Enumeration"""
        self.log("Phase 4: Querying Certificate Transparency logs")
        self.results["phases_executed"]["ct_logs"] = True
        
        ct_domains = set()
        
        domain_targets = [t for t in self.targets if not re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', t)]
        
        if not domain_targets:
            return ct_domains
        
        import urllib.request
        import urllib.parse
        import time
        import socket
        
        self.log(f"Querying CT logs for {len(domain_targets)} domains (sequential mode with retry)")
        
        for idx, target in enumerate(domain_targets, 1):
            self.log(f"Querying CT logs for {target} ({idx}/{len(domain_targets)})")
            
            max_retries = 3
            
            for attempt in range(max_retries):
                try:
                    url = f"https://crt.sh/?q=%.{target}&output=json"
                    self._record_tool("crt.sh", [url])
                    req = urllib.request.Request(url, headers={
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                    })
                    
                    with urllib.request.urlopen(req, timeout=45) as response:
                        data = json.loads(response.read().decode())
                        found_domains = set()
                        for entry in data:
                            name_value = entry.get('name_value', '')
                            for domain in name_value.split('\n'):
                                domain = domain.strip().lower()
                                if domain and not domain.startswith('*'):
                                    found_domains.add(domain)
                        
                        ct_domains.update(found_domains)
                        self.log(f"✓ CT logs for {target}: found {len(found_domains)} domains")
                        break
                        
                except urllib.error.HTTPError as e:
                    if e.code == 429:
                        if attempt < max_retries - 1:
                            wait_time = 10 * (attempt + 1)
                            self.log(f"⚠ Rate limited (429), waiting {wait_time}s...", "WARNING")
                            time.sleep(wait_time)
                        else:
                            self.log(f"✗ Failed to query {target} - rate limited", "ERROR")
                            break
                    else:
                        self.log(f"✗ HTTP Error {e.code} for {target} - skipping", "WARNING")
                        break
                        
                except Exception as e:
                    self.log(f"✗ Unexpected error for {target}: {str(e)} - skipping", "WARNING")
                    break
            
            if idx < len(domain_targets):
                time.sleep(2)
        
        self.results["ct_logs"] = ct_domains
        self.log(f"Found {len(ct_domains)} total domains from CT logs")
        self.save_results("ct_logs", ct_domains)
        return ct_domains
    
    def phase_apex_discovery(self) -> Set[str]:
        """
        Phase 5: APEX Domain Discovery
        Sources:
          - tldextract over all discovered subdomains
          - assetfinder (no --subs-only flag) → grep -iv $domain to keep only non-sub results
          - 365doms.py  → O365 tenant discovery
          - whois       → registrant email domain extraction
        """
        self.log("Phase 5: Discovering APEX/root domains")
        self.results["phases_executed"]["apex"] = True

        apex_domains = set()

        BLOCKLIST = {
            'cloudflare.com', 'cloudflare.net', 'akamai.com', 'akamai.net',
            'fastly.com', 'fastly.net', 'verisign.com', 'identity.digital',
            'icann.org', 'iana.org', 'gandi.net', 'godaddy.com', 'namecheap.com',
            'markmonitor.com', 'whoisguard.com', 'domainsbyproxy.com',
            'gmail.com', 'yahoo.com', 'hotmail.com', 'outlook.com',
            'amazonaws.com', 'googlecloud.com', 'google.com',
            'azure.com', 'microsoft.com', 'digitalocean.com', 'linode.com',
            'windows.net', 'onmicrosoft.com',
        }

        BLOCKLIST_KEYWORDS = [
            'whois', 'gtld', 'registry', 'registrar', 'nameserver',
            'privacy', 'proxy', 'redacted', 'iana', 'icann', 'abuse',
        ]

        # Gather every domain we know about so far
        all_discovered_domains = set()
        for key in ("subdomains", "ct_logs", "reverse_dns_domains",
                    "github_subdomains", "gau_subdomains"):
            all_discovered_domains.update(self.results.get(key, set()))

        if not all_discovered_domains:
            self.log("No subdomain data yet — running subdomain enum first...", "WARNING")
            self.phase_subdomain_enum()
            self.phase_ct_logs()
            all_discovered_domains.update(self.results.get("subdomains", set()))
            all_discovered_domains.update(self.results.get("ct_logs", set()))

        self.log(f"Analyzing {len(all_discovered_domains)} discovered domains for APEX extraction")

        domain_targets = [
            t for t in self.targets
            if not re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}(/\d+)?$', t)
        ]

        # ── tldextract pass over all discovered subdomains ────────────────────
        try:
            import tldextract
            for domain in all_discovered_domains:
                if re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', domain):
                    continue
                extracted = tldextract.extract(domain)
                if extracted.domain and extracted.suffix:
                    apex_domains.add(f"{extracted.domain}.{extracted.suffix}".lower())
        except ImportError:
            for domain in all_discovered_domains:
                if re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', domain):
                    continue
                parts = domain.split('.')
                if len(parts) >= 2:
                    if len(parts) >= 3 and parts[-2] in ['co', 'com', 'org', 'net', 'ac', 'gov', 'edu']:
                        apex = '.'.join(parts[-3:])
                    else:
                        apex = '.'.join(parts[-2:])
                    apex_domains.add(apex.lower())

        # ── assetfinder (no --subs-only) → keep lines that are NOT subdomains ─
        # Bash equivalent: assetfinder $domain | grep -iv $domain
        if self.check_tool("assetfinder"):
            for domain in domain_targets:
                cmd = ["assetfinder", domain]
                self._record_tool("assetfinder (apex)", cmd)
                rc, stdout, _ = self.run_command(cmd, timeout=120)
                if rc == 0 and stdout.strip():
                    for line in stdout.splitlines():
                        line = line.strip().lower()
                        # Keep only lines that do NOT contain the original domain
                        # (these are the related apex/org domains assetfinder finds)
                        if line and '.' in line and domain.lower() not in line:
                            apex_domains.add(line)
            self.log(f"[assetfinder apex] collected additional apex candidates")

        # ── 365doms.py — O365 tenant / related domain discovery ───────────────
        script_365 = Path.home() / "tools" / "365doms.py"
        if script_365.exists():
            for domain in domain_targets:
                self.log(f"[365doms.py] Querying O365 tenants for {domain}...")
                cmd = [sys.executable, str(script_365), "-d", domain]
                self._record_tool("365doms.py", cmd)
                rc, stdout, _ = self.run_command(cmd, timeout=60)
                if rc == 0 and stdout.strip():
                    for line in stdout.splitlines():
                        line = line.strip().lower()
                        if line and '.' in line and not line.startswith('#'):
                            apex_domains.add(line)
                    self.log(f"[365doms.py] Found additional domains for {domain}")
        else:
            self.log("[365doms.py] Script not found at ~/tools/365doms.py — skipping", "WARNING")

        # ── asnmap — find CIDR ranges owned by the org, then extract new apex ──
        # asnmap requires a ProjectDiscovery Cloud Platform API key (PDCP_API_KEY).
        # It returns CIDR blocks associated with the org, which we store for
        # reference; any reverse-DNS names from those CIDRs become apex candidates.
        if self.check_tool("asnmap"):
            pdcp_key = self.pdcp_token or os.environ.get("PDCP_API_KEY", "")
            if not pdcp_key:
                self.log("[asnmap] No PDCP API key — skipping (use --pdcp-token or PDCP_API_KEY env)", "WARNING")
            else:
                self.results["phases_executed"]["asnmap"] = True
                for domain in domain_targets:
                    self.log(f"[asnmap] Mapping ASN ranges for {domain}...")
                    cmd = ["asnmap", "-d", domain, "-silent"]
                    env = os.environ.copy()
                    env["PDCP_API_KEY"] = pdcp_key
                    self._record_tool("asnmap", cmd)
                    try:
                        import subprocess as _sp
                        result = _sp.run(cmd, capture_output=True, text=True, timeout=60, env=env)
                        if result.returncode == 0 and result.stdout.strip():
                            for line in result.stdout.splitlines():
                                cidr = line.strip()
                                if cidr and "/" in cidr:
                                    self.results["asnmap_cidrs"].add(cidr)
                                    self.log(f"[asnmap] Found CIDR: {cidr}")
                    except Exception as e:
                        self.log(f"[asnmap] Error for {domain}: {e}", "WARNING")
                self.log(f"[asnmap] Total CIDRs discovered: {len(self.results['asnmap_cidrs'])}")
        else:
            self.log("[asnmap] Not installed — skipping ASN CIDR mapping", "WARNING")
            self.log("    Install: go install github.com/projectdiscovery/asnmap/cmd/asnmap@latest", "WARNING")

        # ── whois registrant email extraction ─────────────────────────────────
        if self.check_tool("whois") and domain_targets:
            def whois_lookup(target):
                try:
                    whois_cmd = ["whois", target]
                    self._record_tool("whois", whois_cmd)
                    returncode, stdout, _ = self.run_command(whois_cmd)
                    if returncode == 0:
                        found = set()
                        for line in stdout.split('\n'):
                            line_lower = line.lower()
                            if 'registrant email:' in line_lower or 'registrant e-mail:' in line_lower:
                                email_match = re.search(r'@([a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,})', line)
                                if email_match:
                                    found.add(email_match.group(1).lower())
                        return found
                except Exception as e:
                    self.log(f"WHOIS lookup failed for {target}: {e}", "WARNING")
                return set()

            with ThreadPoolExecutor(max_workers=3) as executor:
                futures = {executor.submit(whois_lookup, t): t for t in domain_targets}
                for future in as_completed(futures):
                    apex_domains.update(future.result())

        # ── filter ────────────────────────────────────────────────────────────
        filtered_apex = set()
        target_set_lower = {t.lower() for t in self.targets}
        for domain in apex_domains:
            if domain in BLOCKLIST:
                continue
            if any(kw in domain.lower() for kw in BLOCKLIST_KEYWORDS):
                continue
            if domain in target_set_lower:
                continue
            if '.' in domain and len(domain.split('.')[-1]) >= 2:
                filtered_apex.add(domain)

        self.results["apex_domains"] = filtered_apex
        self.log(f"Discovered {len(filtered_apex)} NEW APEX domains (after filtering)")
        self.save_results("apex_domains", filtered_apex)
        return filtered_apex
    
    def phase_service_fingerprint(self) -> Dict:
        """Phase 6: Service Fingerprinting"""
        self.log("Phase 6: Service fingerprinting")
        self.results["phases_executed"]["services"] = True
        
        services = {}
        
        if not self.results["open_ports"]:
            self.log("No open ports to fingerprint", "WARNING")
            return services
        
        if not self.check_tool("nmap"):
            self.log("nmap not found. Install it first.", "WARNING")
            return services
        
        for ip, ports in list(self.results["open_ports"].items())[:10]:
            port_list = ','.join(map(str, ports[:20]))
            self.log(f"Fingerprinting services on {ip}:{port_list}")
            
            cmd = ["nmap", "-sV", "-p", port_list, ip]
            self._record_tool("nmap", cmd)
            returncode, stdout, _ = self.run_command(cmd, timeout=300)
            
            if returncode == 0:
                services[ip] = stdout
        
        self.results["services"] = services
        self.save_results("services", services)
        return services
    
    def phase_web_crawl(self) -> Set[str]:
        """Phase 7: Web Crawling for additional hosts"""
        self.log("Phase 7: Web crawling for additional discovery")
        self.results["phases_executed"]["web_crawl"] = True
        
        discovered_hosts = set()
        
        web_targets = []
        for ip, ports in self.results["open_ports"].items():
            if 80 in ports:
                web_targets.append(f"http://{ip}")
            if 443 in ports:
                web_targets.append(f"https://{ip}")
        
        for domain in list(self.results["subdomains"])[:20]:
            web_targets.append(f"https://{domain}")
        
        if self.check_tool("hakrawler"):
            for target in web_targets[:10]:
                self.log(f"Crawling {target}")
                hakrawler_cmd = ["hakrawler", "-url", target, "-depth", "2"]
                self._record_tool("hakrawler", hakrawler_cmd)
                returncode, stdout, _ = self.run_command(hakrawler_cmd, timeout=60)

                if returncode == 0:
                    domain_pattern = r'https?://([a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,})'
                    found = re.findall(domain_pattern, stdout)
                    discovered_hosts.update(found)
        else:
            self.log("hakrawler not found. Skipping web crawl.", "WARNING")
        
        self.results["web_crawl"] = discovered_hosts
        self.log(f"Discovered {len(discovered_hosts)} hosts via web crawling")
        self.save_results("web_crawl", discovered_hosts)
        return discovered_hosts

    def phase_bruteforce_subdomains(self) -> Set[str]:
        """
        Bruteforce / permutation phase (run-all only).
        1. Run alterx AND dnsgen over all currently known subdomains to generate
           permutation wordlists.
        2. Resolve the combined wordlist with shuffledns using three resolver tiers
           (extended → standard → trusted).
        3. New valid subdomains are merged back into self.results["subdomains"].

        Bash equivalents:
          alterx -l subdomains.txt -silent | tee alterx_wordlist.txt
          cat subdomains.txt | dnsgen - | tee dnsgen_wordlist.txt
          cat alterx_wordlist.txt dnsgen_wordlist.txt | sort -u > combined_wordlist.txt
          shuffledns -d example.com -list combined_wordlist.txt -r resolvers.txt -mode resolve -silent
        """
        self.log("Phase Bruteforce: Generating permutations and resolving via shuffledns")
        self.results["phases_executed"]["bruteforce"] = True

        known_subs = self.results.get("subdomains", set())
        if not known_subs:
            self.log("[bruteforce] No known subdomains to permute — skipping", "WARNING")
            return set()

        domain_targets = [
            t for t in self.targets
            if not re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}(/\d+)?$', t)
        ]
        if not domain_targets:
            self.log("[bruteforce] No domain targets — skipping", "WARNING")
            return set()

        # ── fetch resolver lists ──────────────────────────────────────────────
        import urllib.request
        resolver_urls = {
            "resolvers-extended.txt": "https://raw.githubusercontent.com/trickest/resolvers/refs/heads/main/resolvers-extended.txt",
            "resolvers.txt":          "https://raw.githubusercontent.com/trickest/resolvers/refs/heads/main/resolvers.txt",
            "resolvers-trusted.txt":  "https://raw.githubusercontent.com/trickest/resolvers/refs/heads/main/resolvers-trusted.txt",
        }
        resolver_files = {}
        for fname, url in resolver_urls.items():
            local = os.path.join(self.output_dir, fname)
            if not os.path.exists(local):
                try:
                    self.log(f"[bruteforce] Fetching {fname}...")
                    urllib.request.urlretrieve(url, local)
                except Exception as e:
                    self.log(f"[bruteforce] Could not fetch {fname}: {e}", "WARNING")
                    local = None
            resolver_files[fname] = local

        # Write known subdomains to temp file
        subs_file = os.path.join(self.output_dir, f"subs_for_bruteforce_{self.timestamp}.txt")
        with open(subs_file, "w") as f:
            f.write("\n".join(sorted(known_subs)))

        # ── generate permutations ─────────────────────────────────────────────
        alterx_out  = os.path.join(self.output_dir, f"alterx_wordlist_{self.timestamp}.txt")
        dnsgen_out  = os.path.join(self.output_dir, f"dnsgen_wordlist_{self.timestamp}.txt")
        combined    = os.path.join(self.output_dir, f"bruteforce_combined_{self.timestamp}.txt")

        generated = set()

        # alterx
        if self.check_tool("alterx"):
            cmd = ["alterx", "-l", subs_file, "-silent", "-o", alterx_out]
            self._record_tool("alterx", cmd)
            self.log(f"[alterx] Generating permutations from {len(known_subs)} subdomains...")
            rc, stdout, _ = self.run_command(cmd, timeout=300)
            if os.path.exists(alterx_out):
                with open(alterx_out) as f:
                    generated.update(l.strip().lower() for l in f if l.strip())
            self.log(f"[alterx] Generated {len(generated)} permutations")
        else:
            self.log("[alterx] Not installed — skipping alterx permutations", "WARNING")

        # dnsgen
        dnsgen_count_before = len(generated)
        if self.check_tool("dnsgen"):
            cmd = ["dnsgen", subs_file, "-o", dnsgen_out]
            self._record_tool("dnsgen", cmd)
            self.log(f"[dnsgen] Generating permutations from {len(known_subs)} subdomains...")
            rc, stdout, _ = self.run_command(cmd, timeout=300)
            if os.path.exists(dnsgen_out):
                with open(dnsgen_out) as f:
                    generated.update(l.strip().lower() for l in f if l.strip())
            elif stdout.strip():
                # Some versions print to stdout
                generated.update(l.strip().lower() for l in stdout.splitlines() if l.strip())
            self.log(f"[dnsgen] Added {len(generated) - dnsgen_count_before} additional permutations")
        else:
            self.log("[dnsgen] Not installed — skipping dnsgen permutations", "WARNING")

        if not generated:
            self.log("[bruteforce] No permutations generated — skipping resolution", "WARNING")
            return set()

        # Write combined wordlist
        with open(combined, "w") as f:
            f.write("\n".join(sorted(generated)))
        self.log(f"[bruteforce] Combined wordlist: {len(generated)} candidates")

        # ── resolve with shuffledns (three resolver tiers) ───────────────────
        if not self.check_tool("shuffledns"):
            self.log("[shuffledns] Not installed — cannot resolve permutations", "WARNING")
            self.log("    Install: go install github.com/projectdiscovery/shuffledns/cmd/shuffledns@latest", "WARNING")
            return set()

        # Ordered from broadest to most trusted; we collect from all three passes
        resolver_tiers = [
            ("resolvers-extended.txt", "extended"),
            ("resolvers.txt",          "standard"),
            ("resolvers-trusted.txt",  "trusted"),
        ]

        new_subs: Set[str] = set()
        for fname, tier_name in resolver_tiers:
            resolver_file = resolver_files.get(fname)
            if not resolver_file or not os.path.exists(resolver_file):
                self.log(f"[shuffledns] Resolver file {fname} not available — skipping {tier_name} tier", "WARNING")
                continue

            shuffledns_out = os.path.join(self.output_dir, f"shuffledns_{tier_name}_{self.timestamp}.txt")

            for domain in domain_targets:
                cmd = [
                    "shuffledns",
                    "-d", domain,
                    "-list", combined,
                    "-r", resolver_file,
                    "-mode", "resolve",
                    "-silent",
                    "-o", shuffledns_out,
                ]
                self._record_tool(f"shuffledns ({tier_name})", cmd)
                self.log(f"[shuffledns] Resolving against {tier_name} resolvers for {domain}...")
                rc, stdout, _ = self.run_command(cmd, timeout=600)

                # Collect from output file
                if os.path.exists(shuffledns_out):
                    with open(shuffledns_out) as f:
                        for line in f:
                            line = line.strip().lower()
                            if line and "." in line and domain in line:
                                new_subs.add(line)
                # Also collect from stdout
                if stdout.strip():
                    for line in stdout.splitlines():
                        line = line.strip().lower()
                        if line and "." in line and domain in line:
                            new_subs.add(line)

        # Only keep genuinely new subdomains
        truly_new = new_subs - known_subs
        if truly_new:
            self.log(f"[bruteforce] ð¯ {len(truly_new)} NEW subdomains discovered via permutation!")
            for s in sorted(truly_new)[:20]:
                self.log(f"  → {s}")
            if len(truly_new) > 20:
                self.log(f"  ... and {len(truly_new) - 20} more")
            self.results["subdomains"] = known_subs.union(truly_new)
            self.save_results("bruteforce_subdomains", truly_new)
        else:
            self.log("[bruteforce] No new subdomains found via permutation")

        self.log(f"[bruteforce] Total subdomains after bruteforce: {len(self.results['subdomains'])}")
        return truly_new
    
    def _format_phase_result(self, phase_key: str):
        if self.results["phases_executed"].get(phase_key, False):
            data = self.results.get(phase_key, None)
            if isinstance(data, dict):
                return len(data)
            elif isinstance(data, (set, list)):
                return len(data)
            return 0
        return "Not Run"

    def _record_tool(self, tool_name: str, cmd: list):
        """Record a tool call with its full command for audit trail"""
        self.results["tools_called"][tool_name] = ' '.join(str(c) for c in cmd)

    def generate_xlsx_report(self):
        """Generate comprehensive XLSX report"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.log("openpyxl not installed. Install with: pip install openpyxl", "ERROR")
            return None
        
        self.log("Generating XLSX report...")
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        new_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        new_font = Font(bold=True, color="006400")
        cdn_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cdn_font = Font(bold=True, color="8B4513")
        github_fill = PatternFill(start_color="C8A2C8", end_color="C8A2C8", fill_type="solid")   # NEW - lilac for GitHub
        github_font = Font(bold=True, color="4B0082")                                              # NEW - dark purple
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        not_run_font = Font(italic=True, color="999999")
        monospace_font = Font(name="Courier New", size=11)

        def _cell(ws, row, col, value, **kwargs):
            """Write a value to a cell, sanitizing strings first so openpyxl
            never receives ANSI codes or other illegal characters."""
            if isinstance(value, str):
                value = self._sanitize(value)
            c = ws.cell(row=row, column=col, value=value)
            for attr, val in kwargs.items():
                setattr(c, attr, val)
            return c
        
        original_targets_lower = {t.lower() for t in self.targets}
        categorized = self.categorize_targets()
        
        # Build domain-IP mapping
        domain_ip_map = {}
        all_domains = (
            self.results['subdomains']
            .union(self.results['ct_logs'])
            .union(self.results.get('reverse_dns_domains', set()))
            .union(self.results.get('github_subdomains', set()))
            .union(self.results.get('shodan_subdomains', set()))
            .union(self.results.get('sublist3r_subdomains', set()))
            .union(set(categorized['domains']))
        )
        
        if self.check_tool("dnsx") and all_domains:
            self.log("Building domain-to-IP mapping using dnsx (batch mode)...")
            temp_domains = os.path.join(self.output_dir, f"temp_report_domains_{self.timestamp}.txt")
            with open(temp_domains, 'w') as f:
                f.write('\n'.join(all_domains))
            
            returncode, stdout, _ = self.run_command(["dnsx", "-l", temp_domains, "-a", "-resp", "-silent"])
            
            if returncode == 0:
                for line in stdout.strip().split('\n'):
                    if line and '[' in line:
                        parts = line.split('[')
                        if len(parts) == 2:
                            domain = parts[0].strip()
                            ip_match = re.search(r'([0-9.]+)', parts[1])
                            if ip_match:
                                domain_ip_map[domain] = ip_match.group(1)
            
            if os.path.exists(temp_domains):
                os.remove(temp_domains)
        
        ip_domain_map = {}
        for domain, ip in domain_ip_map.items():
            if ip not in ip_domain_map:
                ip_domain_map[ip] = []
            ip_domain_map[ip].append(domain)

        cdn_results = self.results.get("cdn_results", {})
        ip_enrichment = self.results.get("ip_enrichment", {})
        github_subdomains = self.results.get("github_subdomains", set())

        # ===== SHEET 1: SUMMARY =====
        ws_summary = wb.create_sheet("Summary")
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 15
        
        summary_data = [
            ["Reconnaissance Summary Report", ""],
            ["", ""],
            ["Timestamp", self.timestamp],
            ["Original Targets Provided", len(self.targets)],
            ["  • IP Addresses (direct)", len(categorized["ips"])],
            ["  • CIDR Ranges", len(categorized["cidrs"])],
            ["  • IPs from CIDR Expansion", len(self.results.get("cidr_expanded_ips", set()))],
            ["  • Domains", len(categorized["domains"])],
            ["", ""],
            ["*** DISCOVERY RESULTS ***", ""],
            ["Domains from Reverse DNS", self._format_phase_result('reverse_dns')],
            ["Total Subdomains Discovered", self._format_phase_result('subdomains')],
            ["  • via GitHub Code Search", self._format_phase_result('github_subdomains')],
            ["  • via gau URL archives",   self._format_phase_result('gau_subdomains')],
            ["  • via sublist3r",           self._format_phase_result('sublist3r_subdomains')],
            ["  • via Shodan (shosubgo)",   self._format_phase_result('shodan_subdomains')],
            ["  • via bruteforce (new)",    self._format_phase_result('bruteforce')],
            ["CT Log Domains Found", self._format_phase_result('ct_logs')],
            ["New APEX Domains Discovered", self._format_phase_result('apex')],
            ["", ""],
            ["*** NETWORK RESULTS ***", ""],
            ["Total Unique IPs Resolved", len(self.results['ip_addresses'])],
            ["ASN CIDRs Discovered (asnmap)", len(self.results.get('asnmap_cidrs', set()))],
            ["CDN/WAF Protected Targets", sum(1 for v in self.results.get('cdn_results',{}).values() if v)],
            ["IPs Skipped (CDN-protected)", sum(1 for ip,v in self.results.get('cdn_results',{}).items() if v and re.match(r'^\d+\.\d+\.\d+\.\d+$', ip))],
            ["IPs Enriched (WHOIS/RDAP)", self._format_phase_result('ip_enrichment')],
            ["Total Hosts with Open Ports", len(self.results['open_ports'])],
            ["Total Open Ports Found", sum(len(ports) for ports in self.results['open_ports'].values())],
            ["", ""],
            ["*** ENUMERATION RESULTS ***", ""],
            ["Service Fingerprinting", self._format_phase_result('services')],
            ["Web Crawl Discoveries", self._format_phase_result('web_crawl')],
            ["Live Web Hosts (httpx)", len(self.results.get("httpx_results", []))],
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, start=1):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            value_cell = ws_summary.cell(row=row_idx, column=2, value=value)
            if value == "Not Run":
                value_cell.font = not_run_font

        # ===== TOOL AUDIT SECTION =====
        TOOL_AUDIT_LABELS = [
            "subfinder",
            "amass",
            "assetfinder (subs-only)",
            "assetfinder (apex)",
            "gau",
            "sublist3r",
            "shosubgo",
            "github-subdomains",
            "asnmap",
            "dnsx (resolution)",
            "dnsx (reverse PTR)",
            "dig (resolution fallback)",
            "dig (reverse PTR fallback)",
            "naabu (top 1000)",
            "naabu (critical ports)",
            "nmap",
            "hakrawler",
            "whois",
            "crt.sh",
            "cdncheck",
            "ipwhois/rdap",
            "httpx",
            "365doms.py",
            "alterx",
            "dnsgen",
            "shuffledns (extended)",
            "shuffledns (standard)",
            "shuffledns (trusted)",
        ]

        tools_called = self.results.get("tools_called", {})

        audit_start_row = len(summary_data) + 2
        ws_summary.cell(row=audit_start_row, column=1, value="")
        ws_summary.cell(row=audit_start_row + 1, column=1, value="*** TOOL AUDIT ***").font = Font(bold=True)
        ws_summary.column_dimensions['B'].width = 80

        current_row = audit_start_row + 2
        for tool_label in TOOL_AUDIT_LABELS:
            command = tools_called.get(tool_label)
            label_cell = ws_summary.cell(row=current_row, column=1, value=tool_label)
            label_cell.font = Font(bold=True)
            value_cell = ws_summary.cell(row=current_row, column=2, value=command if command else "Not Run")
            if command:
                value_cell.font = monospace_font
            else:
                value_cell.font = not_run_font
            current_row += 1

        # ===== SHEET 2: DOMAINS WITH PORTS =====
        ws_domains = wb.create_sheet("Domains")
        ws_domains.column_dimensions['A'].width = 40
        ws_domains.column_dimensions['B'].width = 20
        ws_domains.column_dimensions['C'].width = 20
        ws_domains.column_dimensions['D'].width = 50
        ws_domains.column_dimensions['E'].width = 12
        ws_domains.column_dimensions['F'].width = 15   # NEW - Source column

        headers = ["Domain", "IP Address", "CDN/WAF", "Open Ports", "New Discovery", "Source"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_domains.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        row_idx = 2
        for domain in sorted(all_domains):
            ip = domain_ip_map.get(domain, "Not Resolved")
            cdn_provider = cdn_results.get(domain) or cdn_results.get(ip)
            cdn_str = cdn_provider if cdn_provider else "None"
            ports = self.results['open_ports'].get(ip, [])
            ports_str = ", ".join(map(str, ports)) if ports else ("Skipped (CDN)" if cdn_provider else "No open ports")
            is_new = "YES" if domain.lower() not in original_targets_lower else "NO"

            # Determine source (first match wins — priority order)
            dl = domain.lower()
            if dl in {d.lower() for d in self.results.get('shodan_subdomains', set())}:
                source = "Shodan"
            elif dl in {d.lower() for d in github_subdomains}:
                source = "GitHub"
            elif dl in {d.lower() for d in self.results.get('sublist3r_subdomains', set())}:
                source = "Sublist3r"
            elif dl in {d.lower() for d in self.results.get('gau_subdomains', set())}:
                source = "gau"
            elif dl in {d.lower() for d in self.results.get('ct_logs', set())}:
                source = "CT Logs"
            elif dl in {d.lower() for d in self.results.get('reverse_dns_domains', set())}:
                source = "Reverse DNS"
            elif dl in {d.lower() for d in categorized['domains']}:
                source = "Input"
            else:
                source = "Subfinder/Amass"

            _cell(ws_domains, row_idx, 1, domain, border=border)
            ws_domains.cell(row=row_idx, column=2, value=ip).border = border

            cdn_cell = ws_domains.cell(row=row_idx, column=3, value=cdn_str)
            cdn_cell.border = border
            cdn_cell.alignment = center_align
            if cdn_provider:
                cdn_cell.fill = cdn_fill
                cdn_cell.font = cdn_font

            ws_domains.cell(row=row_idx, column=4, value=ports_str).border = border

            new_cell = ws_domains.cell(row=row_idx, column=5, value=is_new)
            new_cell.border = border
            new_cell.alignment = center_align
            if is_new == "YES":
                new_cell.fill = new_fill
                new_cell.font = new_font

            # Source cell - highlight GitHub findings specially
            source_cell = ws_domains.cell(row=row_idx, column=6, value=source)
            source_cell.border = border
            source_cell.alignment = center_align
            if source == "GitHub":
                source_cell.fill = github_fill
                source_cell.font = github_font

            row_idx += 1

        # ===== SHEET 3: IP ADDRESSES WITH PORTS =====
        ws_ips = wb.create_sheet("IP Addresses")
        ws_ips.column_dimensions['A'].width = 18
        ws_ips.column_dimensions['B'].width = 20
        ws_ips.column_dimensions['C'].width = 12
        ws_ips.column_dimensions['D'].width = 40
        ws_ips.column_dimensions['E'].width = 35
        ws_ips.column_dimensions['F'].width = 30
        ws_ips.column_dimensions['G'].width = 10
        ws_ips.column_dimensions['H'].width = 20
        ws_ips.column_dimensions['I'].width = 12

        headers = [
            "IP Address", "CDN/WAF", "Total Ports", "Open Ports",
            "Associated Domains", "Organization", "Country",
            "Network Range", "New Discovery"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_ips.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        row_idx = 2
        for ip in sorted(self.results['ip_addresses'], key=lambda x: tuple(map(int, x.split('.')))):
            ports = self.results['open_ports'].get(ip, [])
            ports_str = ", ".join(map(str, ports)) if ports else "No open ports"
            domains = ip_domain_map.get(ip, [])
            domains_str = ", ".join(domains) if domains else "Unknown"
            cdn_provider = cdn_results.get(ip)
            cdn_str = cdn_provider if cdn_provider else "None"
            enrich = ip_enrichment.get(ip, {})
            org = enrich.get("organization", "Unknown")
            country = enrich.get("country", "Unknown")
            network_range = enrich.get("cidr", "Unknown")

            is_new = "YES" if (
                ip not in categorized["ips"] and
                ip not in self.results.get("cidr_expanded_ips", set())
            ) else "NO"

            ws_ips.cell(row=row_idx, column=1, value=ip).border = border

            cdn_cell = ws_ips.cell(row=row_idx, column=2, value=cdn_str)
            cdn_cell.border = border
            cdn_cell.alignment = center_align
            if cdn_provider:
                cdn_cell.fill = cdn_fill
                cdn_cell.font = cdn_font

            ws_ips.cell(row=row_idx, column=3, value=len(ports)).border = border
            ws_ips.cell(row=row_idx, column=4, value=ports_str).border = border
            ws_ips.cell(row=row_idx, column=5, value=domains_str).border = border
            ws_ips.cell(row=row_idx, column=6, value=org).border = border
            ws_ips.cell(row=row_idx, column=7, value=country).border = border
            ws_ips.cell(row=row_idx, column=8, value=network_range).border = border

            new_cell = ws_ips.cell(row=row_idx, column=9, value=is_new)
            new_cell.border = border
            new_cell.alignment = center_align
            if is_new == "YES":
                new_cell.fill = new_fill
                new_cell.font = new_font

            row_idx += 1

        # ===== SHEET 4: SERVICES =====
        ws_services = wb.create_sheet("Services")
        ws_services.column_dimensions['A'].width = 20
        ws_services.column_dimensions['B'].width = 10
        ws_services.column_dimensions['C'].width = 60
        
        headers = ["IP Address", "Port", "Service Information"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_services.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row_idx = 2
        if self.results['services']:
            for ip, service_output in self.results['services'].items():
                for line in service_output.split('\n'):
                    if '/tcp' in line or '/udp' in line:
                        ws_services.cell(row=row_idx, column=1, value=ip).border = border
                        ws_services.cell(row=row_idx, column=2, value="").border = border
                        ws_services.cell(row=row_idx, column=3, value=line.strip()).border = border
                        row_idx += 1
        else:
            ws_services.cell(row=2, column=1, value="No service fingerprinting data available")

        # ===== SHEET 5: GITHUB SUBDOMAINS (NEW) =====
        ws_github = wb.create_sheet("GitHub Subdomains")
        ws_github.column_dimensions['A'].width = 45
        ws_github.column_dimensions['B'].width = 20
        ws_github.column_dimensions['C'].width = 30

        headers = ["Subdomain", "IP Address", "Open Ports"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_github.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        if github_subdomains:
            row_idx = 2
            for domain in sorted(github_subdomains):
                ip = domain_ip_map.get(domain, "Not Resolved")
                ports = self.results['open_ports'].get(ip, [])
                ports_str = ", ".join(map(str, ports)) if ports else "No open ports"

                domain_cell = ws_github.cell(row=row_idx, column=1, value=domain)
                domain_cell.border = border
                domain_cell.fill = github_fill
                domain_cell.font = github_font

                ws_github.cell(row=row_idx, column=2, value=ip).border = border
                ws_github.cell(row=row_idx, column=3, value=ports_str).border = border
                row_idx += 1
        else:
            no_data = ws_github.cell(row=2, column=1, value="No GitHub subdomain data — provide --github-token to enable")
            no_data.font = not_run_font

        # ===== SHEET 6: APEX DOMAINS =====
        ws_apex = wb.create_sheet("APEX Domains")
        ws_apex.column_dimensions['A'].width = 38   # APEX Domain
        ws_apex.column_dimensions['B'].width = 12   # Sub Count
        ws_apex.column_dimensions['C'].width = 22   # Discovery Source(s)
        ws_apex.column_dimensions['D'].width = 20   # CDN/WAF
        ws_apex.column_dimensions['E'].width = 18   # IPs
        ws_apex.column_dimensions['F'].width = 72   # Related Subdomains

        apex_headers = ["APEX Domain", "Sub Count", "Discovery Source(s)", "CDN/WAF", "IPs", "Related Subdomains"]
        for col_idx, header in enumerate(apex_headers, start=1):
            cell = ws_apex.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # Build apex → subdomains mapping using tldextract if available
        try:
            import tldextract as _tld
            def _get_apex(d):
                e = _tld.extract(d)
                return f"{e.domain}.{e.suffix}".lower() if e.domain and e.suffix else None
        except ImportError:
            def _get_apex(d):
                parts = d.split('.')
                if len(parts) < 2:
                    return None
                if len(parts) >= 3 and parts[-2] in ['co','com','org','net','ac','gov','edu']:
                    return '.'.join(parts[-3:])
                return '.'.join(parts[-2:])

        apex_subdomain_map = {}   # apex → [subdomains]
        apex_ip_map       = {}   # apex → set of IPs
        apex_cdn_map      = {}   # apex → set of CDN providers

        all_known_subs = (
            self.results.get('subdomains', set())
            .union(self.results.get('ct_logs', set()))
            .union(self.results.get('reverse_dns_domains', set()))
            .union(self.results.get('github_subdomains', set()))
            .union(self.results.get('gau_subdomains', set()))
        )

        for subdomain in all_known_subs:
            apex = _get_apex(subdomain)
            if not apex:
                continue
            apex_subdomain_map.setdefault(apex, []).append(subdomain)
            # Resolve IP from domain_ip_map built earlier
            ip = domain_ip_map.get(subdomain)
            if ip:
                apex_ip_map.setdefault(apex, set()).add(ip)
                cdn = cdn_results.get(subdomain) or cdn_results.get(ip)
                if cdn:
                    apex_cdn_map.setdefault(apex, set()).add(cdn)

        # Determine discovery source per apex domain
        # We attribute based on which result sets contributed subdomains under this apex
        def _apex_sources(apex):
            sources = []
            sub_set = set(apex_subdomain_map.get(apex, []))
            if sub_set & self.results.get('subdomains', set()):
                sources.append("subfinder/amass")
            if sub_set & self.results.get('github_subdomains', set()):
                sources.append("GitHub")
            if sub_set & self.results.get('gau_subdomains', set()):
                sources.append("gau")
            if sub_set & self.results.get('ct_logs', set()):
                sources.append("CT Logs")
            if sub_set & self.results.get('reverse_dns_domains', set()):
                sources.append("Reverse DNS")
            # Also mark if apex itself came from assetfinder/365doms (no subdomains under it yet)
            if not sources:
                sources.append("assetfinder/365doms")
            return ", ".join(sources)

        # Style variations
        apex_source_fill  = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        wrap_align        = Alignment(wrap_text=True, vertical="top")
        count_align       = Alignment(horizontal="center", vertical="top")

        row_idx = 2
        for apex in sorted(self.results['apex_domains']):
            related       = sorted(apex_subdomain_map.get(apex, []))
            sub_count     = len(related)
            related_str   = "\n".join(related) if related else "None found"
            sources_str   = _apex_sources(apex)
            ips_str       = ", ".join(sorted(apex_ip_map.get(apex, set()))) or "Not resolved"
            cdn_str       = ", ".join(sorted(apex_cdn_map.get(apex, set()))) or "None"

            # APEX Domain cell — highlighted green (new discovery)
            apex_cell = ws_apex.cell(row=row_idx, column=1, value=apex)
            apex_cell.border    = border
            apex_cell.fill      = new_fill
            apex_cell.font      = new_font
            apex_cell.alignment = Alignment(vertical="top")

            # Sub Count
            count_cell = ws_apex.cell(row=row_idx, column=2, value=sub_count)
            count_cell.border    = border
            count_cell.alignment = count_align

            # Discovery Source(s)
            src_cell = ws_apex.cell(row=row_idx, column=3, value=sources_str)
            src_cell.border    = border
            src_cell.fill      = apex_source_fill
            src_cell.alignment = Alignment(vertical="top")

            # CDN/WAF
            cdn_cell = ws_apex.cell(row=row_idx, column=4, value=cdn_str)
            cdn_cell.border    = border
            cdn_cell.alignment = Alignment(vertical="top")
            if cdn_str != "None":
                cdn_cell.fill = cdn_fill
                cdn_cell.font = cdn_font

            # IPs
            ip_cell = ws_apex.cell(row=row_idx, column=5, value=ips_str)
            ip_cell.border    = border
            ip_cell.alignment = Alignment(vertical="top")

            # Related Subdomains — one per line, wrapped
            rel_cell = ws_apex.cell(row=row_idx, column=6, value=related_str)
            rel_cell.border    = border
            rel_cell.alignment = wrap_align

            # Row height: taller rows for apex domains with many subs (cap at 200pt)
            ws_apex.row_dimensions[row_idx].height = min(15 * max(sub_count, 1), 200)

            row_idx += 1

        # ===== SHEET 7: LIVE WEB HOSTS (httpx probe results) =====
        ws_httpx = wb.create_sheet("Live Web Hosts")
        ws_httpx.column_dimensions['A'].width = 45   # URL
        ws_httpx.column_dimensions['B'].width = 10   # Status
        ws_httpx.column_dimensions['C'].width = 40   # Title
        ws_httpx.column_dimensions['D'].width = 22   # Server
        ws_httpx.column_dimensions['E'].width = 30   # Tech
        ws_httpx.column_dimensions['F'].width = 16   # IP
        ws_httpx.column_dimensions['G'].width = 18   # CDN

        httpx_headers = ["URL", "Status", "Title", "Server", "Tech Stack", "IP", "CDN/WAF"]
        for col_idx, header in enumerate(httpx_headers, start=1):
            cell = ws_httpx.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        httpx_results = self.results.get("httpx_results", [])

        # Status code fill helpers
        def _status_fill(code):
            try:
                code = int(code)
            except (TypeError, ValueError):
                return None
            if 200 <= code < 300:
                return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
            if 300 <= code < 400:
                return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # yellow
            if 400 <= code < 500:
                return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # red
            if 500 <= code < 600:
                return PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # dark red
            return None

        if httpx_results:
            # Sort by status code then URL
            httpx_results_sorted = sorted(
                httpx_results,
                key=lambda r: (int(r.get("status_code") or 0), r.get("url", ""))
            )
            row_idx = 2
            for r in httpx_results_sorted:
                status     = r.get("status_code", "")
                sf         = _status_fill(status)

                url_cell = ws_httpx.cell(row=row_idx, column=1, value=r.get("url", ""))
                url_cell.border = border

                sc_cell = ws_httpx.cell(row=row_idx, column=2, value=status)
                sc_cell.border    = border
                sc_cell.alignment = center_align
                if sf:
                    sc_cell.fill = sf

                ws_httpx.cell(row=row_idx, column=3, value=self._sanitize(r.get("title", "") or "")).border  = border
                ws_httpx.cell(row=row_idx, column=4, value=r.get("server", "")).border  = border
                ws_httpx.cell(row=row_idx, column=5, value=r.get("tech", "")).border    = border
                ws_httpx.cell(row=row_idx, column=6, value=r.get("ip", "")).border      = border

                cdn_val  = r.get("cdn", "") or "None"
                cdn_cell = ws_httpx.cell(row=row_idx, column=7, value=cdn_val)
                cdn_cell.border = border
                if cdn_val and cdn_val != "None":
                    cdn_cell.fill = cdn_fill
                    cdn_cell.font = cdn_font

                row_idx += 1
        else:
            no_data = ws_httpx.cell(row=2, column=1,
                value="No live web hosts found — httpx probe returned no results")
            no_data.font = not_run_font
            ws_httpx.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

        # ===== SHEET 8: ASN RANGES (asnmap — review only, NOT scanned) =====
        ws_asn = wb.create_sheet("ASN Ranges")
        ws_asn.column_dimensions['A'].width = 22   # CIDR
        ws_asn.column_dimensions['B'].width = 30   # Org (from ip_enrichment if available)
        ws_asn.column_dimensions['C'].width = 10   # Country
        ws_asn.column_dimensions['D'].width = 14   # ASN
        ws_asn.column_dimensions['E'].width = 50   # Notes

        asn_headers = ["CIDR Range", "Organization", "Country", "ASN", "Notes"]
        for col_idx, header in enumerate(asn_headers, start=1):
            cell = ws_asn.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # Warning banner row
        warn_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        warn_font = Font(bold=True, color="FFFFFF", size=11)
        warn_cell = ws_asn.cell(row=2, column=1,
            value="⚠  THESE RANGES ARE FOR REVIEW ONLY — NOT SCANNED. Confirm scope with client before any testing.")
        warn_cell.fill = warn_font_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        warn_cell.font = Font(bold=True, color="FFFFFF")
        warn_cell.alignment = Alignment(horizontal="left")
        ws_asn.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

        asnmap_cidrs = self.results.get("asnmap_cidrs", set())
        if asnmap_cidrs:
            # Build a quick CIDR→enrichment lookup from ip_enrichment if we have it
            # (ip_enrichment is keyed by IP, not CIDR — best-effort match on asn_cidr field)
            cidr_enrich = {}   # cidr → {org, country, asn}
            for ip_data in ip_enrichment.values():
                cidr_key = ip_data.get("cidr", "")
                if cidr_key and cidr_key not in cidr_enrich:
                    cidr_enrich[cidr_key] = {
                        "org":     ip_data.get("organization", ""),
                        "country": ip_data.get("country", ""),
                        "asn":     ip_data.get("asn", ""),
                    }

            asn_data_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow

            row_idx = 3
            for cidr in sorted(asnmap_cidrs):
                enrich = cidr_enrich.get(cidr, {})
                org     = enrich.get("org", "—")
                country = enrich.get("country", "—")
                asn     = enrich.get("asn", "—")

                cidr_cell = ws_asn.cell(row=row_idx, column=1, value=cidr)
                cidr_cell.border = border
                cidr_cell.fill   = asn_data_fill
                cidr_cell.font   = Font(bold=True)

                ws_asn.cell(row=row_idx, column=2, value=org).border     = border
                ws_asn.cell(row=row_idx, column=3, value=country).border = border
                ws_asn.cell(row=row_idx, column=4, value=asn).border     = border

                notes_cell = ws_asn.cell(row=row_idx, column=5,
                    value="Confirm with client — scope this range before testing")
                notes_cell.border = border
                notes_cell.font   = Font(italic=True, color="666666")

                row_idx += 1
        else:
            no_asn = ws_asn.cell(row=3, column=1,
                value="No ASN ranges discovered — provide --pdcp-token to enable asnmap")
            no_asn.font = not_run_font
            ws_asn.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)

        # ===== SHEET 9: SCOPE CONFIRMATION =====
        ws_scope = wb.create_sheet("Scope Confirmation")
        ws_scope.column_dimensions['A'].width = 80
        
        scope_text = [
            "*** SCOPE CONFIRMATION FOR CLIENT ***",
            "",
            "Based on our reconnaissance, we have identified the following assets in scope for the penetration test:",
            "",
            "DOMAINS:",
        ]
        
        for domain in sorted(all_domains):
            is_new = " [NEWLY DISCOVERED]" if domain.lower() not in original_targets_lower else ""
            github_tag = " [GITHUB]" if domain.lower() in {d.lower() for d in github_subdomains} else ""
            scope_text.append(f"  • {domain}{is_new}{github_tag}")
        
        scope_text.extend(["", "IP ADDRESSES:"])
        
        for ip in sorted(self.results['ip_addresses'], key=lambda x: tuple(map(int, x.split('.')))):
            is_new_ip = ip not in categorized["ips"]
            new_marker = " [NEWLY DISCOVERED]" if is_new_ip else ""
            scope_text.append(f"  • {ip}{new_marker}")
        
        scope_text.extend(["", "NEWLY DISCOVERED APEX DOMAINS:"])
        
        if self.results['apex_domains']:
            for apex in sorted(self.results['apex_domains']):
                scope_text.append(f"  • {apex}")
        else:
            scope_text.append("  • None")

        if github_subdomains:
            scope_text.extend(["", f"GITHUB SUBDOMAIN DISCOVERIES ({len(github_subdomains)} found):"])
            for domain in sorted(github_subdomains):
                scope_text.append(f"  • {domain}")

        shodan_subs = self.results.get("shodan_subdomains", set())
        if shodan_subs:
            scope_text.extend(["", f"SHODAN SUBDOMAIN DISCOVERIES ({len(shodan_subs)} found):"])
            for domain in sorted(shodan_subs):
                scope_text.append(f"  • {domain}")

        asnmap_cidrs = self.results.get("asnmap_cidrs", set())
        if asnmap_cidrs:
            scope_text.extend(["", f"ASN CIDR RANGES DISCOVERED via asnmap ({len(asnmap_cidrs)} ranges):"])
            for cidr in sorted(asnmap_cidrs):
                scope_text.append(f"  • {cidr}")
        
        scope_text.extend([
            "",
            f"TOTAL ASSETS: {len(all_domains)} domains, {len(self.results['ip_addresses'])} IPs",
            f"TOTAL OPEN PORTS: {sum(len(ports) for ports in self.results['open_ports'].values())}",
            "",
            "Please confirm if all discovered assets should be included in the scope of testing.",
        ])
        
        for row_idx, text in enumerate(scope_text, start=1):
            cell = ws_scope.cell(row=row_idx, column=1, value=text)
            if "***" in text or "NEWLY DISCOVERED" in text:
                cell.font = Font(bold=True)
            if "[NEWLY DISCOVERED]" in text or "[GITHUB]" in text:
                cell.font = Font(bold=True, color="006400")
        
        xlsx_file = os.path.join(self.output_dir, f"recon_report_{self.timestamp}.xlsx")
        wb.save(xlsx_file)
        self.log(f"XLSX report saved: {xlsx_file}")
        
        return xlsx_file
    
    def run_pm_scan(self):
        """
        PM Mode (--pm): Fast scope-mapping for pre-meeting / SOW prep.

        Runs:
          ✅ Subdomain enumeration   (all passive tools in parallel)
          ✅ CT log query
          ✅ Reverse DNS             (if IPs provided)
          ✅ APEX domain discovery   (tldextract + assetfinder + 365doms + asnmap)
          ✅ DNS resolution          (resolve subs → IPs)
          ✅ IP enrichment           (WHOIS/RDAP — ASN, org, country)

        Skips:
          ❌ Port scanning  (naabu)
          ❌ CDN check      (not needed without port scan)
          ❌ Service fingerprinting (nmap)
          ❌ Web crawling   (hakrawler)
          ❌ Bruteforce     (alterx + dnsgen + shuffledns)

        Typical runtime: 3-8 minutes depending on number of targets and
        how many subdomain tools are installed / have API keys.
        Bruteforce alone can add 15-45 minutes — that is why it is excluded.
        Web crawl adds up to 10 minutes — also excluded.
        """
        self.log("=" * 60)
        self.log("[PM MODE] Scope mapping — subdomains, APEX, ASN, IPs only")
        self.log("  Skipped: port scan, service fingerprint, crawl, bruteforce")
        self.log("=" * 60)

        categorized = self.categorize_targets()
        ip_targets   = set(categorized["ips"])
        domain_targets = set(categorized["domains"])

        self.log(f"\nTarget Analysis:")
        self.log(f"  • IP Addresses: {len(ip_targets)}")
        self.log(f"  • CIDR Ranges:  {len(categorized['cidrs'])}")
        self.log(f"  • Domains:      {len(domain_targets)}")

        all_domains = set()
        all_ips     = ip_targets.copy()

        # CIDR expansion (for reverse DNS only — CIDRs are NOT port scanned)
        if categorized["cidrs"]:
            self.log(f"\n[CIDR EXPANSION] Expanding {len(categorized['cidrs'])} range(s)...")
            expanded = self.expand_cidr_targets(categorized["cidrs"])
            all_ips.update(expanded)

        # Reverse DNS on any provided IPs
        if all_ips:
            self.log("\n[REVERSE DNS] Discovering domains from provided IPs...")
            self.results["ip_addresses"] = all_ips
            reverse_domains = self.phase_reverse_dns(all_ips)
            all_domains.update(reverse_domains)

        # Passive subdomain enumeration + CT logs in parallel
        if domain_targets or all_domains:
            self.log("\n[DISCOVERY] Subdomain enumeration + CT logs (parallel)...")
            domains_to_enumerate = domain_targets.union(all_domains)
            original_targets = self.targets
            self.targets = list(domains_to_enumerate)

            with ThreadPoolExecutor(max_workers=2) as executor:
                sub_future = executor.submit(self.phase_subdomain_enum)
                ct_future  = executor.submit(self.phase_ct_logs)
                subdomains = sub_future.result()
                ct_domains = ct_future.result()

            self.targets = original_targets
            all_domains.update(subdomains)
            all_domains.update(ct_domains)
            all_domains.update(domain_targets)

            # APEX discovery (includes asnmap for ASN ranges, assetfinder, 365doms)
            self.log("\n[APEX] Discovering root/org domains and ASN ranges...")
            apex_domains = self.phase_apex_discovery()
            if apex_domains:
                all_domains.update(apex_domains)

        # DNS resolution — get IPs for all discovered domains
        if all_domains:
            self.log("\n[RESOLUTION] Resolving all domains to IPs...")
            resolved_ips = self.phase_dns_resolution(all_domains)
            all_ips.update(resolved_ips)

        # IP enrichment — ASN / org / country (no port scan)
        if all_ips:
            self.log("\n[ENRICHMENT] WHOIS/RDAP enrichment for ASN, org, country...")
            self.results["ip_addresses"] = all_ips
            self.phase_ip_enrichment(all_ips)

        # HTTP probing — phase builds the full domain set internally from all sources
        self.log("\n[HTTP PROBE] Checking live web endpoints (ports 80,443,8080,8443,8888,9443)...")
        self.phase_httpx_probe()

        self.log("\n[REPORTING] Generating report...")
        self.deduplicate_results()
        self.generate_report()

        # Print a clean scope summary to stdout for quick copy-paste
        live_web = len(self.results.get("httpx_results", []))
        self.log("=" * 60)
        self.log("PM SCAN COMPLETE — scope summary:")
        self.log(f"  Subdomains found : {len(self.results['subdomains'])}")
        self.log(f"  Unique IPs       : {len(self.results['ip_addresses'])}")
        self.log(f"  APEX domains     : {len(self.results['apex_domains'])}")
        self.log(f"  ASN CIDR ranges  : {len(self.results.get('asnmap_cidrs', set()))}")
        self.log(f"  Live web hosts   : {live_web}")
        self.log("  (No ports scanned — share report with client for scope confirmation)")
        self.log("=" * 60)

    def run_default_scan(self):
        """Execute default reconnaissance (skips services and web crawl for speed)"""
        self.log("=" * 60)
        self.log("[DEFAULT SCAN MODE] Running core reconnaissance")
        self.log("(Excludes: service fingerprinting, web crawling)")
        self.log("Use --run-all for comprehensive scan")
        self.log("=" * 60)

        categorized = self.categorize_targets()
        ip_targets = set(categorized["ips"])
        domain_targets = set(categorized["domains"])

        self.log(f"\nTarget Analysis:")
        self.log(f"  • IP Addresses: {len(ip_targets)}")
        self.log(f"  • CIDR Ranges:  {len(categorized['cidrs'])}")
        self.log(f"  • Domains:      {len(domain_targets)}")

        all_domains = set()
        all_ips = ip_targets.copy()

        if categorized["cidrs"]:
            self.log(f"\n[CIDR EXPANSION] Expanding {len(categorized['cidrs'])} CIDR range(s)...")
            expanded = self.expand_cidr_targets(categorized["cidrs"])
            all_ips.update(expanded)

        if ip_targets:
            self.log("\n[IP DISCOVERY PHASE] Processing provided IP addresses...")
            self.results["ip_addresses"] = ip_targets
            reverse_domains = self.phase_reverse_dns(ip_targets)
            all_domains.update(reverse_domains)

        if domain_targets or all_domains:
            self.log("\n[DISCOVERY PHASE] Finding subdomains and domains...")
            domains_to_enumerate = domain_targets.union(all_domains)
            original_targets = self.targets
            self.targets = list(domains_to_enumerate)

            # phase_subdomain_enum runs ALL tools in parallel internally
            with ThreadPoolExecutor(max_workers=2) as executor:
                subdomain_future = executor.submit(self.phase_subdomain_enum)
                ct_future = executor.submit(self.phase_ct_logs)
                subdomains = subdomain_future.result()
                ct_domains = ct_future.result()

            self.targets = original_targets

            all_domains.update(subdomains)
            all_domains.update(ct_domains)
            all_domains.update(domain_targets)

            apex_domains = self.phase_apex_discovery()
            if apex_domains:
                all_domains.update(apex_domains)

        if all_domains:
            self.log("\n[RESOLUTION PHASE] Resolving all domains to IPs...")
            resolved_ips = self.phase_dns_resolution(all_domains)
            all_ips.update(resolved_ips)

        if all_ips or all_domains:
            self.log("\n[CDN CHECK PHASE] Identifying CDN/WAF protected targets...")
            self.phase_cdn_check(all_ips.union(all_domains))

        if all_ips:
            self.log("\n[SCANNING PHASE] Port scanning all IPs...")
            self.results["ip_addresses"] = all_ips
            self.phase_port_scan(all_ips)

        if self.results["ip_addresses"]:
            self.log("\n[ENRICHMENT PHASE] Enriching IPs with WHOIS/RDAP data...")
            self.phase_ip_enrichment(self.results["ip_addresses"])

        self.log("\n[HTTP PROBE PHASE] Checking live web endpoints (80,443,8080,8443,8888,9443)...")
        self.phase_httpx_probe()

        self.log("\n[REPORTING PHASE] Generating final report...")
        self.deduplicate_results()
        self.generate_report()
    
    def run_all_phases(self):
        """Execute all reconnaissance phases including services and web crawl"""
        self.log("=" * 60)
        self.log("[FULL SCAN MODE] Running all reconnaissance phases")
        self.log("(Includes: service fingerprinting, web crawling, GitHub subdomain search)")
        self.log("=" * 60)

        categorized = self.categorize_targets()
        ip_targets = set(categorized["ips"])
        domain_targets = set(categorized["domains"])

        all_domains = set()
        all_ips = ip_targets.copy()

        if categorized["cidrs"]:
            expanded = self.expand_cidr_targets(categorized["cidrs"])
            all_ips.update(expanded)

        if ip_targets:
            self.results["ip_addresses"] = ip_targets
            reverse_domains = self.phase_reverse_dns(ip_targets)
            all_domains.update(reverse_domains)

        if domain_targets or all_domains:
            domains_to_enumerate = domain_targets.union(all_domains)
            original_targets = self.targets
            self.targets = list(domains_to_enumerate)

            # phase_subdomain_enum runs ALL tools in parallel internally
            with ThreadPoolExecutor(max_workers=2) as executor:
                subdomain_future = executor.submit(self.phase_subdomain_enum)
                ct_future = executor.submit(self.phase_ct_logs)
                subdomains = subdomain_future.result()
                ct_domains = ct_future.result()

            self.targets = original_targets

            all_domains.update(subdomains)
            all_domains.update(ct_domains)
            all_domains.update(domain_targets)

            apex_domains = self.phase_apex_discovery()
            if apex_domains:
                all_domains.update(apex_domains)

        if all_domains:
            resolved_ips = self.phase_dns_resolution(all_domains)
            all_ips.update(resolved_ips)

        if all_ips or all_domains:
            self.phase_cdn_check(all_ips.union(all_domains))

        if all_ips:
            self.results["ip_addresses"] = all_ips
            self.phase_port_scan(all_ips)

        if self.results["ip_addresses"]:
            self.phase_ip_enrichment(self.results["ip_addresses"])

        if self.results["open_ports"]:
            with ThreadPoolExecutor(max_workers=2) as executor:
                service_future = executor.submit(self.phase_service_fingerprint)
                crawl_future = executor.submit(self.phase_web_crawl)
                service_future.result()
                crawl_future.result()

        # ── BRUTEFORCE PHASE (run-all only) ───────────────────────────────────
        # Runs AFTER passive enum so we permute the real discovered subdomains,
        # then resolves with shuffledns + three resolver tiers.
        # New finds are merged back into subdomains before DNS resolution reruns.
        self.log("\n[BRUTEFORCE PHASE] Permutation + shuffledns resolution...")
        new_from_bruteforce = self.phase_bruteforce_subdomains()
        if new_from_bruteforce:
            self.log(f"Re-resolving {len(new_from_bruteforce)} bruteforce discoveries...")
            extra_ips = self.phase_dns_resolution(new_from_bruteforce)
            all_ips.update(extra_ips)
            self.results["ip_addresses"].update(extra_ips)

        self.log("\n[HTTP PROBE PHASE] Checking live web endpoints (80,443,8080,8443,8888,9443)...")
        # Called with no argument — phase builds from all known sources including bruteforce
        self.phase_httpx_probe()

        self.deduplicate_results()
        self.generate_report()
    
    def generate_report(self):
        """Generate summary report (both TXT and XLSX)"""
        categorized = self.categorize_targets()
        ip_targets = categorized["ips"]
        domain_targets = categorized["domains"]
        
        report_file = os.path.join(self.output_dir, f"recon_summary_{self.timestamp}.txt")
        
        with open(report_file, 'w') as f:
            f.write("=" * 60 + "\n")
            f.write("RECONNAISSANCE SUMMARY REPORT\n")
            f.write("=" * 60 + "\n\n")
            f.write(f"Timestamp: {self.timestamp}\n")
            f.write(f"Initial Targets: {len(self.targets)}\n")
            f.write(f"  • IP Addresses Provided: {len(ip_targets)}\n")
            f.write(f"  • Domains Provided: {len(domain_targets)}\n\n")
            
            f.write("=== DISCOVERY RESULTS ===\n")
            f.write(f"Domains from Reverse DNS: {len(self.results.get('reverse_dns_domains', set()))}\n")
            f.write(f"Subdomains Discovered: {len(self.results['subdomains'])}\n")
            f.write(f"  • via GitHub Code Search: {len(self.results.get('github_subdomains', set()))}\n")
            f.write(f"IP Addresses Resolved: {len(self.results['ip_addresses'])}\n")
            f.write(f"Hosts with Open Ports: {len(self.results['open_ports'])}\n")
            f.write(f"CT Log Domains: {len(self.results['ct_logs'])}\n")
            f.write(f"APEX Domains: {len(self.results['apex_domains'])}\n")
            f.write(f"Web Crawl Discoveries: {len(self.results['web_crawl'])}\n")
            
            f.write("\n" + "=" * 60 + "\n")
            f.write("All detailed results saved in: " + self.output_dir + "\n")
            f.write("=" * 60 + "\n")
        
        self.log(f"Summary report generated: {report_file}")
        print(f"\n{open(report_file).read()}")
        
        xlsx_file = self.generate_xlsx_report()
        
        if xlsx_file:
            self.log("=" * 60)
            self.log("REPORTS GENERATED:")
            self.log(f"  • Text Summary: {report_file}")
            self.log(f"  • Excel Report: {xlsx_file}")
            self.log("=" * 60)


def main():
    dep_manager = DependencyManager()
    dep_manager.check_and_install_all()

    parser = argparse.ArgumentParser(
        description="External Network Recon Framework for Authorized Penetration Testing",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Default fast scan (skips services/crawl)
  %(prog)s -t example.com
  %(prog)s -f targets.txt

  # Full scan with GitHub subdomain search
  %(prog)s -t example.com --run-all --github-token ghp_xxxxxxxxxxxx

  # Default scan with GitHub token
  %(prog)s -f targets.txt --github-token ghp_xxxxxxxxxxxx

  # GitHub subdomain search only
  %(prog)s -t example.com --github-subdomains --github-token ghp_xxxxxxxxxxxx

  # CIDR ranges
  %(prog)s -t 192.168.1.0/24 example.com 10.0.0.1

  # Run specific phases
  %(prog)s -t example.com --subdomains --dns --ports
  %(prog)s -t 192.168.1.0/24 --cdn-check --ip-enrichment

GitHub Token:
  Get a token at https://github.com/settings/tokens
  Only 'public_repo' read scope is needed (no write access required)
  Token is REDACTED in all audit logs and Excel reports for security

Scan Modes:
  Default (no flags):  Core scan - subdomains, DNS, ports, CDN, enrichment (~3-5 min)
  --pm:                PM/pre-meeting mode - subs, APEX, ASN, IPs, no port scan (~3-8 min)
  --run-all:           Full scan - all phases including bruteforce (~5-10 min + bruteforce)

Bruteforce timing note (--run-all only):
  Small target  (<50 known subs)  : ~3-5 min
  Medium target (~200 known subs) : ~10-20 min
  Large target  (500+ known subs) : ~30-60 min
  Use --pm or default scan to skip bruteforce entirely.

Platform Support:
  - macOS:   Full auto-install
  - Linux:   Full auto-install
  - WSL2:    Full auto-install (recommended for Windows users)
  - Windows: Limited features (use WSL2 for best experience)
        """
    )

    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument('-t', '--targets', nargs='+',
                             help='Target domains, IPs, or CIDR ranges (can be mixed)')
    input_group.add_argument('-f', '--file',
                             help='File containing targets (one per line, # for comments)')

    parser.add_argument('-o', '--output', default='recon_output',
                        help='Output directory (default: recon_output)')
    parser.add_argument('--run-all', action='store_true',
                        help='Run ALL phases including service enumeration and web crawling')
    parser.add_argument('--pm', action='store_true',
                        help=('PM / pre-meeting mode: subdomains, APEX, ASN ranges, IPs, WHOIS enrichment. '
                              'No port scan, no nmap, no crawl, no bruteforce. ~3-8 min.'))
    parser.add_argument('--github-token', metavar='TOKEN',
                        help='GitHub personal access token for github-subdomains (public_repo read scope)')
    parser.add_argument('--shodan-token', metavar='TOKEN',
                        help='Shodan API key for shosubgo subdomain discovery')
    parser.add_argument('--pdcp-token', metavar='TOKEN',
                        help='ProjectDiscovery Cloud Platform API key for asnmap (also reads PDCP_API_KEY env var)')

    # Individual phase flags
    parser.add_argument('--subdomains',          action='store_true', help='Run subdomain enumeration')
    parser.add_argument('--github-subdomains',   action='store_true', help='Run GitHub subdomain search (requires --github-token)')
    parser.add_argument('--dns',                 action='store_true', help='Run DNS resolution')
    parser.add_argument('--ports',               action='store_true', help='Run port scanning')
    parser.add_argument('--ct-logs',             action='store_true', help='Run CT log enumeration')
    parser.add_argument('--apex',                action='store_true', help='Run APEX domain discovery')
    parser.add_argument('--services',            action='store_true', help='Run service fingerprinting')
    parser.add_argument('--crawl',               action='store_true', help='Run web crawling')
    parser.add_argument('--reverse-dns',         action='store_true', help='Run reverse DNS')
    parser.add_argument('--cdn-check',           action='store_true', help='Run CDN/WAF check')
    parser.add_argument('--ip-enrichment',       action='store_true', help='Run IP WHOIS/RDAP enrichment')
    parser.add_argument('--bruteforce',          action='store_true', help='Run permutation bruteforce (alterx+dnsgen+shuffledns). Implies subdomains must exist first')

    args = parser.parse_args()

    if args.targets:
        targets = args.targets
    else:
        with open(args.file, 'r') as f:
            targets = [
                line.strip() for line in f
                if line.strip() and not line.strip().startswith('#')
            ]

    if not targets:
        print("Error: No targets provided")
        sys.exit(1)

    recon = ReconFramework(targets, args.output,
                           github_token=args.github_token,
                           shodan_token=args.shodan_token,
                           pdcp_token=args.pdcp_token)

    phase_flags = [
        args.subdomains, args.github_subdomains, args.dns, args.ports, args.ct_logs,
        args.apex, args.services, args.crawl, args.reverse_dns,
        args.cdn_check, args.ip_enrichment, args.bruteforce
    ]

    if args.pm:
        recon.run_pm_scan()

    elif args.run_all:
        recon.run_all_phases()

    elif any(phase_flags):
        categorized = recon.categorize_targets()

        all_ips_from_cidr = set()
        if categorized["cidrs"]:
            all_ips_from_cidr = recon.expand_cidr_targets(categorized["cidrs"])

        if args.reverse_dns:
            all_ips = set(categorized["ips"]).union(all_ips_from_cidr)
            if all_ips:
                recon.results["ip_addresses"] = all_ips
                recon.phase_reverse_dns(all_ips)
            else:
                recon.log("No IP addresses to perform reverse DNS on", "WARNING")

        if args.subdomains:
            recon.phase_subdomain_enum()

        if args.github_subdomains:
            # Runs the full unified enum (all tools); GitHub results land in
            # results["github_subdomains"] automatically when a token is present
            recon.phase_subdomain_enum()

        if args.dns:
            recon.phase_dns_resolution(set(targets))

        if args.ct_logs:
            recon.phase_ct_logs()

        if args.cdn_check:
            all_ips = set(categorized["ips"]).union(all_ips_from_cidr)
            recon.phase_cdn_check(all_ips.union(set(categorized["domains"])))

        if args.ports:
            recon.phase_port_scan(set(targets))

        if args.ip_enrichment:
            if not recon.results["ip_addresses"]:
                recon.phase_dns_resolution(set(targets))
            recon.phase_ip_enrichment(recon.results["ip_addresses"])

        if args.services:
            if not recon.results["open_ports"]:
                recon.log("Service fingerprinting needs port data. Running port scan first...")
                if not recon.results["ip_addresses"]:
                    recon.phase_dns_resolution(set(targets))
                recon.phase_port_scan()
            recon.phase_service_fingerprint()

        if args.crawl:
            if not recon.results["open_ports"]:
                recon.log("Web crawling needs port data. Running port scan first...")
                if not recon.results["ip_addresses"]:
                    recon.phase_dns_resolution(set(targets))
                recon.phase_port_scan()
            recon.phase_web_crawl()

        if args.apex:
            recon.phase_apex_discovery()

        if args.bruteforce:
            if not recon.results["subdomains"]:
                recon.log("Bruteforce needs subdomains first — running subdomain enum...")
                recon.phase_subdomain_enum()
            new_subs = recon.phase_bruteforce_subdomains()
            if new_subs:
                recon.phase_dns_resolution(new_subs)

        recon.deduplicate_results()
        recon.generate_report()

    else:
        recon.run_default_scan()


if __name__ == "__main__":
    main()
